"""
WearCheck ARC - Calibration Certificate Template Generator
===========================================================
Creates a structured Excel workbook for calibration data entry.
Sheets: Config, Input, Calibration_Data, Reference_Standards, Audit_Log

ISO 17025 aligned: traceability, uncertainty, repeatability.
"""

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule
from datetime import datetime
import os

# ── Styling Constants ──────────────────────────────────────────────────────

WEARCHECK_RED = "C2040B"
WEARCHECK_BLACK = "1A1A1A"
WEARCHECK_GREY = "4D4D4D"
WEARCHECK_LTGREY = "E6E6E6"
HEADER_FILL = PatternFill(start_color=WEARCHECK_RED, end_color=WEARCHECK_RED, fill_type="solid")
ACCENT_FILL = PatternFill(start_color=WEARCHECK_BLACK, end_color=WEARCHECK_BLACK, fill_type="solid")
LIGHT_GREY = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
SECTION_BG = PatternFill(start_color=WEARCHECK_LTGREY, end_color=WEARCHECK_LTGREY, fill_type="solid")
PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
TITLE_FONT = Font(bold=True, size=14, color=WEARCHECK_RED)
SECTION_FONT = Font(bold=True, size=11, color=WEARCHECK_RED)
BOLD_FONT = Font(bold=True, size=10)
NORMAL_FONT = Font(size=10)
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def apply_border_range(ws, min_row, max_row, min_col, max_col):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER


def style_header_row(ws, row, min_col, max_col, fill=HEADER_FILL, font=HEADER_FONT):
    for col in range(min_col, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill
        cell.font = font
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def create_config_sheet(wb):
    """Config sheet: dropdown lists, tolerances, controlled values."""
    ws = wb.create_sheet("Config")

    # ── Equipment Types ──
    ws["A1"] = "Equipment Types"
    ws["A1"].font = SECTION_FONT
    equipment = [
        "Spectrometer", "Particle Counter", "Viscometer",
        "Moisture Analyser", "Ferrograph", "Titrator",
        "pH Meter", "Conductivity Meter", "Other"
    ]
    for i, eq in enumerate(equipment, 2):
        ws[f"A{i}"] = eq

    # ── Detector Types ──
    ws["B1"] = "Detector Types"
    ws["B1"].font = SECTION_FONT
    detectors = [
        "Survey Probe", "Inclinometer", "Gyroscope",
        "Magnetometer", "Accelerometer", "N/A"
    ]
    for i, d in enumerate(detectors, 2):
        ws[f"B{i}"] = d

    # ── Calibration Procedures ──
    ws["C1"] = "Procedure IDs"
    ws["C1"].font = SECTION_FONT
    procedures = [
        "WC-CAL-001 Spectrometer Calibration",
        "WC-CAL-002 Particle Counter Calibration",
        "WC-CAL-003 Viscometer Calibration",
        "WC-CAL-004 Detector Calibration",
        "WC-CAL-005 Inclinometer Calibration",
        "WC-CAL-006 Gyroscope Calibration",
        "WC-CAL-007 General Instrument Calibration",
    ]
    for i, p in enumerate(procedures, 2):
        ws[f"C{i}"] = p

    # ── Tolerances ──
    ws["E1"] = "Tolerance Table"
    ws["E1"].font = SECTION_FONT
    ws["E2"] = "Parameter"
    ws["F2"] = "Unit"
    ws["G2"] = "Tolerance (+/-)"
    style_header_row(ws, 2, 5, 7, fill=ACCENT_FILL, font=Font(bold=True, size=10, color="FFFFFF"))

    tolerances = [
        ("Detector Response", "%", 2.0),
        ("Inclinometer", "degrees", 0.5),
        ("Gyroscope", "degrees", 1.0),
        ("Temperature", "°C", 2.0),
        ("Humidity", "%RH", 5.0),
        ("Voltage", "V", 0.05),
        ("Current", "mA", 0.01),
    ]
    for i, (param, unit, tol) in enumerate(tolerances, 3):
        ws[f"E{i}"] = param
        ws[f"F{i}"] = unit
        ws[f"G{i}"] = tol
        for col in range(5, 8):
            ws.cell(row=i, column=col).border = THIN_BORDER
            ws.cell(row=i, column=col).alignment = CENTER

    # ── Technicians ──
    ws["I1"] = "Technicians"
    ws["I1"].font = SECTION_FONT
    techs = ["Tech 1", "Tech 2", "Tech 3"]
    for i, t in enumerate(techs, 2):
        ws[f"I{i}"] = t

    # ── Reviewers ──
    ws["J1"] = "Reviewers"
    ws["J1"].font = SECTION_FONT
    reviewers = ["Reviewer 1", "Reviewer 2"]
    for i, r in enumerate(reviewers, 2):
        ws[f"J{i}"] = r

    # ── Approvers ──
    ws["K1"] = "Approvers"
    ws["K1"].font = SECTION_FONT
    approvers = ["Approver 1", "Approver 2"]
    for i, a in enumerate(approvers, 2):
        ws[f"K{i}"] = a

    # ── Certificate Number Format ──
    ws["M1"] = "Cert Number Prefix"
    ws["M1"].font = SECTION_FONT
    ws["M2"] = "WC-ARC-CAL-"
    ws["M3"] = "Next Sequence"
    ws["M3"].font = BOLD_FONT
    ws["M4"] = 1

    # Column widths
    for col_letter in ["A", "B", "C", "E", "F", "G", "I", "J", "K", "M"]:
        ws.column_dimensions[col_letter].width = 22

    ws.sheet_state = "hidden"  # Hide config from technicians
    return ws


def create_input_sheet(wb):
    """Input sheet: certificate metadata, equipment details, environment.
    
    Sequential layout — all sections flow top to bottom to avoid
    merged-cell collisions between left and right columns.
    """
    ws = wb.create_sheet("Certificate_Input")

    # Column widths
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 4
    ws.column_dimensions["E"].width = 28
    ws.column_dimensions["F"].width = 30

    # ── Title ──
    ws.merge_cells("B2:F2")
    ws["B2"] = "WEARCHECK ARC — CALIBRATION CERTIFICATE INPUT"
    ws["B2"].font = TITLE_FONT
    ws["B2"].alignment = CENTER

    ws.merge_cells("B3:F3")
    ws["B3"] = "Complete all fields below. Yellow cells require input."
    ws["B3"].font = Font(italic=True, size=9, color="666666")
    ws["B3"].alignment = CENTER

    row = 5
    input_fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")

    def add_section(title, start_row):
        ws.merge_cells(f"B{start_row}:F{start_row}")
        ws[f"B{start_row}"] = title
        ws[f"B{start_row}"].font = SECTION_FONT
        ws[f"B{start_row}"].fill = SECTION_BG
        for c in ["C", "D", "E", "F"]:
            ws[f"{c}{start_row}"].fill = SECTION_BG
        return start_row + 1

    def add_field(label, row, col_label=2, col_value=3):
        ws.cell(row=row, column=col_label, value=label).font = BOLD_FONT
        ws.cell(row=row, column=col_label).alignment = LEFT
        ws.cell(row=row, column=col_label).border = THIN_BORDER
        ws.cell(row=row, column=col_value).fill = input_fill
        ws.cell(row=row, column=col_value).border = THIN_BORDER
        ws.cell(row=row, column=col_value).alignment = LEFT
        return row + 1

    def add_field_pair(label_l, label_r, row):
        """Add a left (B/C) and right (E/F) field pair on the same row."""
        ws.cell(row=row, column=2, value=label_l).font = BOLD_FONT
        ws.cell(row=row, column=2).alignment = LEFT
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=3).fill = input_fill
        ws.cell(row=row, column=3).border = THIN_BORDER
        ws.cell(row=row, column=3).alignment = LEFT
        ws.cell(row=row, column=5, value=label_r).font = BOLD_FONT
        ws.cell(row=row, column=5).alignment = LEFT
        ws.cell(row=row, column=5).border = THIN_BORDER
        ws.cell(row=row, column=6).fill = input_fill
        ws.cell(row=row, column=6).border = THIN_BORDER
        ws.cell(row=row, column=6).alignment = LEFT
        return row + 1

    # ══════════════════════════════════════════════════════════════════════
    # Section: Certificate Info
    # ══════════════════════════════════════════════════════════════════════
    row = add_section("CERTIFICATE INFORMATION", row)

    # Certificate Number (auto-generated)
    ws.cell(row=row, column=2, value="Certificate Number").font = BOLD_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=3).border = THIN_BORDER
    ws.cell(row=row, column=3).fill = input_fill
    ws.cell(row=row, column=3, value='=Config!M2&TEXT(Config!M4,"0000")')
    ws.cell(row=row, column=3).font = BOLD_FONT
    row += 1

    row = add_field("Date of Issue", row)
    ws.cell(row=row - 1, column=3).number_format = "YYYY-MM-DD"

    row = add_field("Calibration Date", row)
    ws.cell(row=row - 1, column=3).number_format = "YYYY-MM-DD"

    row = add_field("Next Calibration Due", row)
    ws.cell(row=row - 1, column=3).number_format = "YYYY-MM-DD"

    row = add_field("Calibration Procedure", row)
    proc_dv = DataValidation(type="list", formula1="Config!$C$2:$C$8", allow_blank=True)
    proc_dv.error = "Select a valid procedure from the list"
    proc_dv.errorTitle = "Invalid Procedure"
    ws.add_data_validation(proc_dv)
    proc_dv.add(ws.cell(row=row - 1, column=3))

    row += 1  # spacing

    # ══════════════════════════════════════════════════════════════════════
    # Section: Equipment Under Test
    # ══════════════════════════════════════════════════════════════════════
    row = add_section("EQUIPMENT UNDER TEST (EUT)", row)
    equip_name_row = row
    row = add_field("Equipment Name", row)
    row = add_field("Manufacturer", row)
    row = add_field("Model Number", row)
    row = add_field("Serial Number", row)
    row = add_field("Asset Number", row)
    row = add_field("Location / Department", row)

    equip_dv = DataValidation(type="list", formula1="Config!$A$2:$A$10", allow_blank=True)
    equip_dv.error = "Select a valid equipment type"
    ws.add_data_validation(equip_dv)
    equip_dv.add(ws.cell(row=equip_name_row, column=3))

    row += 1  # spacing

    # ══════════════════════════════════════════════════════════════════════
    # Section: Specifications & Tolerances
    # ══════════════════════════════════════════════════════════════════════
    row = add_section("SPECIFICATIONS & TOLERANCES", row)
    det_tol_row = row
    row = add_field("Detector Tolerance (+/-)", row)
    inc_tol_row = row
    row = add_field("Inclinometer Tolerance (+/-)", row)
    gyro_tol_row = row
    row = add_field("Gyroscope Tolerance (+/-)", row)

    row += 1  # spacing

    # ══════════════════════════════════════════════════════════════════════
    # Section: Environmental Conditions
    # ══════════════════════════════════════════════════════════════════════
    row = add_section("ENVIRONMENTAL CONDITIONS", row)
    temp_row = row
    row = add_field("Temperature (°C)", row)
    hum_row = row
    row = add_field("Humidity (%RH)", row)
    row = add_field("Atmospheric Pressure (kPa)", row)

    temp_dv = DataValidation(type="decimal", operator="between", formula1=10, formula2=40)
    temp_dv.error = "Temperature must be between 10°C and 40°C"
    temp_dv.errorTitle = "Temperature Out of Range"
    ws.add_data_validation(temp_dv)
    temp_dv.add(ws.cell(row=temp_row, column=3))

    hum_dv = DataValidation(type="decimal", operator="between", formula1=10, formula2=90)
    hum_dv.error = "Humidity must be between 10% and 90%"
    ws.add_data_validation(hum_dv)
    hum_dv.add(ws.cell(row=hum_row, column=3))

    row += 1  # spacing

    # ══════════════════════════════════════════════════════════════════════
    # Section: Calibration Software
    # ══════════════════════════════════════════════════════════════════════
    row = add_section("CALIBRATION SOFTWARE", row)
    row = add_field("Software Manufacturer", row)
    row = add_field("Software Name", row)
    row = add_field("Software Version", row)

    row += 1  # spacing

    # ══════════════════════════════════════════════════════════════════════
    # Section: Mounting Conditions
    # ══════════════════════════════════════════════════════════════════════
    row = add_section("MOUNTING CONDITIONS & CONSIDERATIONS", row)
    row = add_field("Mounting Torque", row)
    row = add_field("Lubrication Used", row)
    row = add_field("Mounting Orientation", row)
    row = add_field("Cable Routing", row)
    row = add_field("Reference Level", row)
    row = add_field("gn Value", row)

    row += 1  # spacing

    # ══════════════════════════════════════════════════════════════════════
    # Section: Authorisation — Technician
    # ══════════════════════════════════════════════════════════════════════
    row = add_section("AUTHORISATION — CALIBRATED BY (TECHNICIAN)", row)
    tech_name_row = row
    row = add_field_pair("Tech Name", "Tech Surname", row)
    row = add_field_pair("Tech Title", "Tech Date Signed", row)
    ws.cell(row=row - 1, column=6).number_format = "YYYY-MM-DD"
    row = add_field("Tech Signature", row)
    ws.cell(row=row - 1, column=3).fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")

    row += 1  # spacing

    # Section: Authorisation — Reviewer
    row = add_section("AUTHORISATION — REVIEWED BY", row)
    rev_name_row = row
    row = add_field_pair("Reviewer Name", "Reviewer Surname", row)
    row = add_field_pair("Reviewer Title", "Reviewer Date Signed", row)
    ws.cell(row=row - 1, column=6).number_format = "YYYY-MM-DD"
    row = add_field("Reviewer Signature", row)
    ws.cell(row=row - 1, column=3).fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")

    row += 1  # spacing

    # Section: Authorisation — Approver
    row = add_section("AUTHORISATION — APPROVED BY", row)
    appr_name_row = row
    row = add_field_pair("Approver Name", "Approver Surname", row)
    row = add_field_pair("Approver Title", "Approver Date Signed", row)
    ws.cell(row=row - 1, column=6).number_format = "YYYY-MM-DD"
    row = add_field("Approver Signature", row)
    ws.cell(row=row - 1, column=3).fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")

    row += 1  # spacing

    # ══════════════════════════════════════════════════════════════════════
    # Section: Traceability & Uncertainty Statements
    # ══════════════════════════════════════════════════════════════════════
    row = add_section("TRACEABILITY & UNCERTAINTY", row)
    ws.merge_cells(f"B{row}:F{row}")
    ws.cell(row=row, column=2).value = (
        "Traceability: The measurements reported herein are traceable to national/international "
        "standards through an unbroken chain of calibrations. Reference standards used "
        "are documented in the Reference_Standards sheet."
    )
    ws.cell(row=row, column=2).font = Font(italic=True, size=9)
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 30
    row += 1

    ws.merge_cells(f"B{row}:F{row}")
    ws.cell(row=row, column=2).value = (
        "Uncertainty: The reported expanded uncertainty is based on a standard uncertainty multiplied "
        "by a coverage factor k=2, providing a level of confidence of approximately 95%. "
        "The uncertainty evaluation is in accordance with the Guide to the Expression of "
        "Uncertainty in Measurement (GUM)."
    )
    ws.cell(row=row, column=2).font = Font(italic=True, size=9)
    ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 30
    row += 2  # spacing

    # ══════════════════════════════════════════════════════════════════════
    # Section: Results Summary (auto-calculated)
    # ══════════════════════════════════════════════════════════════════════
    row = add_section("RESULTS SUMMARY (AUTO-CALCULATED)", row)

    row = add_field("Pre-Calibration Result", row)
    ws.cell(row=row - 1, column=3).value = (
        '=IF(Calibration_Data!H3="","N/A",Calibration_Data!H3)'
    )
    row = add_field("Post-Calibration Result", row)
    ws.cell(row=row - 1, column=3).value = (
        '=IF(Calibration_Data!H21="","N/A",Calibration_Data!H21)'
    )
    row = add_field("Inclinometer Result", row)
    ws.cell(row=row - 1, column=3).value = (
        '=IF(Calibration_Data!H39="","N/A",Calibration_Data!H39)'
    )
    row = add_field("Gyroscope Result", row)
    ws.cell(row=row - 1, column=3).value = (
        '=IF(Calibration_Data!H57="","N/A",Calibration_Data!H57)'
    )

    result_start = row - 4
    row = add_field("OVERALL RESULT", row)
    ws.cell(row=row - 1, column=3).value = (
        f'=IF(COUNTIF(C{result_start}:C{result_start+3},"FAIL")>0,"FAIL",'
        f'IF(COUNTIF(C{result_start}:C{result_start+3},"PASS")>0,"PASS","N/A"))'
    )
    ws.cell(row=row - 1, column=3).font = Font(bold=True, size=12)

    # Conditional formatting for PASS/FAIL
    pass_rule = CellIsRule(operator="equal", formula=['"PASS"'], fill=PASS_FILL)
    fail_rule = CellIsRule(operator="equal", formula=['"FAIL"'], fill=FAIL_FILL)
    result_range = f"C{result_start}:C{row - 1}"
    ws.conditional_formatting.add(result_range, pass_rule)
    ws.conditional_formatting.add(result_range, fail_rule)

    # Print setup
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    return ws, det_tol_row, inc_tol_row, gyro_tol_row


def create_calibration_data_sheet(wb, det_tol_row, inc_tol_row, gyro_tol_row):
    """Calibration data sheet with auto-calculated deviations and conformance."""
    ws = wb.create_sheet("Calibration_Data")

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 16
    ws.column_dimensions["H"].width = 18

    input_fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")
    NUM_POINTS = 15  # number of calibration points per section

    def create_cal_table(title, start_row, tol_ref, section_label):
        """Create a calibration data table with formulas.

        tol_ref: cell reference for tolerance, e.g. "Certificate_Input!C22"
        Returns the row after the table ends.
        """
        # Section title
        ws.merge_cells(f"B{start_row}:H{start_row}")
        ws[f"B{start_row}"] = title
        ws[f"B{start_row}"].font = SECTION_FONT
        ws[f"B{start_row}"].fill = SECTION_BG
        for c in range(3, 9):
            ws.cell(row=start_row, column=c).fill = SECTION_BG

        # Headers
        hr = start_row + 1
        headers = ["#", "Set Point", "Actual Reading", "Deviation", "Tolerance (+/-)", "Conformance"]
        cols = [2, 3, 4, 5, 6, 7]
        for col, header in zip(cols, headers):
            cell = ws.cell(row=hr, column=col, value=header)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = CENTER
            cell.border = THIN_BORDER

        # Section result header
        ws.cell(row=hr, column=8, value="Section Result").fill = HEADER_FILL
        ws.cell(row=hr, column=8).font = HEADER_FONT
        ws.cell(row=hr, column=8).alignment = CENTER
        ws.cell(row=hr, column=8).border = THIN_BORDER

        # Data rows
        data_start = hr + 1
        for i in range(NUM_POINTS):
            r = data_start + i
            # Row number
            ws.cell(row=r, column=2, value=i + 1).alignment = CENTER
            ws.cell(row=r, column=2).border = THIN_BORDER
            ws.cell(row=r, column=2).fill = LIGHT_GREY

            # Set Point (manual entry)
            ws.cell(row=r, column=3).fill = input_fill
            ws.cell(row=r, column=3).border = THIN_BORDER
            ws.cell(row=r, column=3).alignment = CENTER

            # Actual Reading (manual entry)
            ws.cell(row=r, column=4).fill = input_fill
            ws.cell(row=r, column=4).border = THIN_BORDER
            ws.cell(row=r, column=4).alignment = CENTER

            # Deviation = Actual - Set Point (auto)
            dev_formula = f'=IF(AND(C{r}<>"",D{r}<>""),D{r}-C{r},"")'
            ws.cell(row=r, column=5, value=dev_formula).alignment = CENTER
            ws.cell(row=r, column=5).border = THIN_BORDER
            ws.cell(row=r, column=5).number_format = "0.000"

            # Tolerance (linked from input)
            tol_formula = f'=IF(C{r}<>"",{tol_ref},"")'
            ws.cell(row=r, column=6, value=tol_formula).alignment = CENTER
            ws.cell(row=r, column=6).border = THIN_BORDER
            ws.cell(row=r, column=6).number_format = "0.000"

            # Conformance: PASS/FAIL (auto)
            conf_formula = (
                f'=IF(E{r}="","",IF(ABS(E{r})<=F{r},"PASS","FAIL"))'
            )
            ws.cell(row=r, column=7, value=conf_formula).alignment = CENTER
            ws.cell(row=r, column=7).border = THIN_BORDER
            ws.cell(row=r, column=7).font = BOLD_FONT

        data_end = data_start + NUM_POINTS - 1

        # Section result formula (in the first data row, column H)
        result_formula = (
            f'=IF(COUNTBLANK(G{data_start}:G{data_end})={NUM_POINTS},"",IF(COUNTIF(G{data_start}:G{data_end},"FAIL")>0,"FAIL","PASS"))'
        )
        ws.cell(row=data_start, column=8, value=result_formula)
        ws.cell(row=data_start, column=8).font = Font(bold=True, size=12)
        ws.cell(row=data_start, column=8).alignment = CENTER
        ws.cell(row=data_start, column=8).border = THIN_BORDER

        # Merge the result cell across rows
        if NUM_POINTS > 1:
            ws.merge_cells(f"H{data_start}:H{data_end}")

        # Conditional formatting for conformance column
        pass_rule = CellIsRule(operator="equal", formula=['"PASS"'], fill=PASS_FILL)
        fail_rule = CellIsRule(operator="equal", formula=['"FAIL"'], fill=FAIL_FILL)
        conf_range = f"G{data_start}:G{data_end}"
        ws.conditional_formatting.add(conf_range, pass_rule)
        ws.conditional_formatting.add(conf_range, fail_rule)

        # Conditional formatting for section result
        result_range = f"H{data_start}:H{data_start}"
        ws.conditional_formatting.add(result_range, pass_rule)
        ws.conditional_formatting.add(result_range, fail_rule)

        return data_end + 2  # leave a gap

    # Tolerance references from Certificate_Input sheet
    det_tol_ref = f"Certificate_Input!C{det_tol_row}"
    inc_tol_ref = f"Certificate_Input!C{inc_tol_row}"
    gyro_tol_ref = f"Certificate_Input!C{gyro_tol_row}"

    # ── Table 1: Before Calibration (Detector) ──
    next_row = create_cal_table(
        "TABLE 1: BEFORE CALIBRATION — DETECTOR RESPONSE",
        1, det_tol_ref, "pre_detector"
    )

    # ── Table 2: After Calibration (Detector) ──
    next_row = create_cal_table(
        "TABLE 2: AFTER CALIBRATION — DETECTOR RESPONSE",
        next_row, det_tol_ref, "post_detector"
    )

    # ── Table 3: Inclinometer Check ──
    next_row = create_cal_table(
        "TABLE 3: INCLINOMETER CHECK",
        next_row, inc_tol_ref, "inclinometer"
    )

    # ── Table 4: Gyroscope Check ──
    next_row = create_cal_table(
        "TABLE 4: GYROSCOPE CHECK",
        next_row, gyro_tol_ref, "gyroscope"
    )

    # Print setup
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0

    return ws


def create_reference_standards_sheet(wb):
    """Reference standards and traceability data."""
    ws = wb.create_sheet("Reference_Standards")

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 28
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 20
    ws.column_dimensions["G"].width = 20
    ws.column_dimensions["H"].width = 20

    input_fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")

    # Title
    ws.merge_cells("B1:H1")
    ws["B1"] = "REFERENCE STANDARDS & TRACEABILITY"
    ws["B1"].font = TITLE_FONT
    ws["B1"].alignment = CENTER

    # Headers
    headers = ["#", "Reference Standard", "Manufacturer", "Serial Number",
               "Certificate Number", "Calibration Due", "Traceability"]
    for i, h in enumerate(headers, 2):
        cell = ws.cell(row=3, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 10 rows for reference standards
    for row in range(4, 14):
        ws.cell(row=row, column=2, value=row - 3).alignment = CENTER
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).fill = LIGHT_GREY
        for col in range(3, 9):
            ws.cell(row=row, column=col).fill = input_fill
            ws.cell(row=row, column=col).border = THIN_BORDER
            ws.cell(row=row, column=col).alignment = CENTER

    # Date validation for calibration due column
    date_dv = DataValidation(type="date", allow_blank=True)
    date_dv.error = "Enter a valid date"
    ws.add_data_validation(date_dv)
    for row in range(4, 14):
        date_dv.add(ws.cell(row=row, column=7))
        ws.cell(row=row, column=7).number_format = "YYYY-MM-DD"

    return ws


def create_audit_log_sheet(wb):
    """Audit log for tracking changes and reviews."""
    ws = wb.create_sheet("Audit_Log")

    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 18
    ws.column_dimensions["F"].width = 40
    ws.column_dimensions["G"].width = 20

    # Title
    ws.merge_cells("B1:G1")
    ws["B1"] = "AUDIT LOG — CERTIFICATE CHANGE HISTORY"
    ws["B1"].font = TITLE_FONT
    ws["B1"].alignment = CENTER

    ws.merge_cells("B2:G2")
    ws["B2"] = "Record all changes, reviews, and approvals below."
    ws["B2"].font = Font(italic=True, size=9, color="666666")
    ws["B2"].alignment = CENTER

    # Headers
    headers = ["#", "Date", "Action", "Performed By", "Description", "Verified By"]
    for i, h in enumerate(headers, 2):
        cell = ws.cell(row=4, column=i, value=h)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    input_fill = PatternFill(start_color="FFFFF0", end_color="FFFFF0", fill_type="solid")

    # Pre-populate first entry
    ws.cell(row=5, column=2, value=1).alignment = CENTER
    ws.cell(row=5, column=2).border = THIN_BORDER
    ws.cell(row=5, column=3, value=datetime.now().strftime("%Y-%m-%d"))
    ws.cell(row=5, column=3).border = THIN_BORDER
    ws.cell(row=5, column=4, value="Certificate Created")
    ws.cell(row=5, column=4).border = THIN_BORDER
    ws.cell(row=5, column=5, value="System")
    ws.cell(row=5, column=5).border = THIN_BORDER
    ws.cell(row=5, column=6, value="Initial certificate template generated")
    ws.cell(row=5, column=6).border = THIN_BORDER
    ws.cell(row=5, column=7).border = THIN_BORDER

    # 50 blank rows for future entries
    for row in range(6, 56):
        ws.cell(row=row, column=2, value=row - 4).alignment = CENTER
        ws.cell(row=row, column=2).border = THIN_BORDER
        ws.cell(row=row, column=2).fill = LIGHT_GREY
        for col in range(3, 8):
            ws.cell(row=row, column=col).fill = input_fill
            ws.cell(row=row, column=col).border = THIN_BORDER

    # Protect certain columns from accidental editing
    actions_dv = DataValidation(
        type="list",
        formula1='"Certificate Created,Data Entry,Calibration Performed,Review Completed,Approval Granted,Amendment,Reissue,Voided"',
        allow_blank=True
    )
    actions_dv.error = "Select a valid action type"
    ws.add_data_validation(actions_dv)
    for row in range(5, 56):
        actions_dv.add(ws.cell(row=row, column=4))

    return ws


def build_workbook():
    """Main function: assemble the complete workbook."""
    wb = openpyxl.Workbook()

    # Remove default sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    # Create sheets in order
    config_ws = create_config_sheet(wb)
    input_ws, det_tol_row, inc_tol_row, gyro_tol_row = create_input_sheet(wb)
    cal_ws = create_calibration_data_sheet(wb, det_tol_row, inc_tol_row, gyro_tol_row)
    ref_ws = create_reference_standards_sheet(wb)
    audit_ws = create_audit_log_sheet(wb)

    # Set Certificate_Input as the active sheet
    wb.active = wb.sheetnames.index("Certificate_Input")

    # Save
    output_dir = os.path.dirname(os.path.abspath(__file__))
    output_path = os.path.join(output_dir, "WearCheck_ARC_Calibration_Template.xlsx")
    wb.save(output_path)
    print(f"Template created: {output_path}")
    return output_path


if __name__ == "__main__":
    build_workbook()
