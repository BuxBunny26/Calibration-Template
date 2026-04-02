"""
WearCheck ARC — Calibration Certificate Cover Page Generator
=============================================================
Reads a filled PVC Report Generation Workbook (.xlsm saved from .xltm)
and generates a branded single-page PDF cover page to attach to the
Modal Shop PVC calibration report.

Uses the same design as TEST_Certificate.pdf — logo top-left header,
light-grey section bars, field tables, signature block, branded footer.

Usage:
    python generate_cover_page.py <filled_workbook.xlsm>

Output:
    CoverPage_<SensorModel>_<SerialNumber>_<Date>.pdf
"""

import os
import sys
from datetime import datetime, date, time as dt_time

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image, HRFlowable
)
from reportlab.platypus.flowables import KeepTogether

import openpyxl

# ── Brand Colors (identical to generate_certificate.py) ──────────────────

WEARCHECK_RED = colors.HexColor("#C2040B")
WEARCHECK_BLACK = colors.HexColor("#1A1A1A")
WEARCHECK_GREY = colors.HexColor("#4D4D4D")
WEARCHECK_LTGREY = colors.HexColor("#E6E6E6")
LIGHT_GREY = colors.HexColor("#F2F2F2")
PASS_GREEN = colors.HexColor("#C6EFCE")
DARK_GREEN = colors.HexColor("#006100")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH = os.path.join(SCRIPT_DIR, "WearCheck Logo.png")
PAGE_W, PAGE_H = A4
MARGIN = 15 * mm

# Values that indicate empty/error cells from formulas
_BLANK = {None, "", "#N/A", "#VALUE!", "#REF!", "#DIV/0!", "0", "0.0"}


# ── Styles (matching TEST_Certificate.pdf) ────────────────────────────────

def get_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="CertTitle", fontName="Helvetica-Bold", fontSize=16,
        textColor=WEARCHECK_BLACK, alignment=TA_CENTER, spaceAfter=2 * mm,
    ))
    styles.add(ParagraphStyle(
        name="SectionHead", fontName="Helvetica-Bold", fontSize=9,
        textColor=WEARCHECK_BLACK, alignment=TA_LEFT,
        leftIndent=2 * mm, spaceBefore=1.5 * mm, spaceAfter=0.5 * mm,
    ))
    styles.add(ParagraphStyle(
        name="SubSectionHead", fontName="Helvetica-Bold", fontSize=8,
        textColor=colors.HexColor("#4D4D4D"), alignment=TA_LEFT,
        leftIndent=4 * mm, spaceBefore=1 * mm, spaceAfter=0.5 * mm,
    ))
    styles.add(ParagraphStyle(
        name="FieldLabel", fontName="Helvetica-Bold", fontSize=7.5,
        textColor=WEARCHECK_BLACK, alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        name="FieldValue", fontName="Helvetica", fontSize=8,
        alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        name="SmallItalic", fontName="Helvetica-Oblique", fontSize=6.5,
        textColor=colors.grey, alignment=TA_JUSTIFY, spaceBefore=1 * mm,
    ))
    styles.add(ParagraphStyle(
        name="Footer", fontName="Helvetica", fontSize=6,
        textColor=colors.grey, alignment=TA_CENTER,
    ))
    return styles


# ── PVC Workbook Reader ──────────────────────────────────────────────────

def _val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    s = str(v).strip()
    if s in _BLANK:
        return ""
    return s


def _date_val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, dt_time):
        if v == dt_time(0, 0):
            return ""
        return v.strftime("%H:%M")
    if v is None:
        return ""
    return str(v).strip()


def detect_cert_type(wb):
    """Detect which certificate type has data by checking for sensor info."""
    # Check FRCert first (most common)
    if "FRRaw" in wb.sheetnames:
        ws = wb["FRRaw"]
        # If FRRaw has data in the sensor identification cells
        if _val(ws, 5, 1) or _val(ws, 5, 2):  # A5=Model, B5=Serial
            return "FR"

    if "LINRaw" in wb.sheetnames:
        ws = wb["LINRaw"]
        if _val(ws, 5, 1) or _val(ws, 5, 2):
            return "LIN"

    # Check cert sheets directly for computed values
    for name, tag in [("FRCert", "FR"), ("LINCert", "LIN"),
                      ("SLINCert", "SLIN"), ("Vib Analyzer Cert", "VIB")]:
        if name in wb.sheetnames:
            ws = wb[name]
            # Check if model/serial cells have values
            if tag == "FR":
                if _val(ws, 5, 3) or _val(ws, 6, 3):
                    return tag
            elif tag in ("LIN", "SLIN"):
                if _val(ws, 5, 5) or _val(ws, 6, 5):
                    return tag
            elif tag == "VIB":
                if _val(ws, 5, 3) or _val(ws, 6, 3):
                    return tag

    # Also check Vib Analyzer Data sheet (Cert sheet may have uncached formulas)
    if "Vib Analyzer Data" in wb.sheetnames:
        ws = wb["Vib Analyzer Data"]
        if _val(ws, 5, 3) or _val(ws, 6, 3):
            return "VIB"

    return "FR"  # Default


def read_fr_info(wb):
    """Extract info from Frequency Response certificate sheets."""
    cert = wb["FRCert"]
    raw = wb["FRRaw"] if "FRRaw" in wb.sheetnames else None
    data = wb["FRData"] if "FRData" in wb.sheetnames else None

    info = {
        "cert_type": "Frequency Response Calibration",
        "model": _val(cert, 5, 3),
        "serial": _val(cert, 6, 3),
        "manufacturer": _val(cert, 7, 3),
        "id_number": _val(cert, 8, 3),
        "description": _val(cert, 9, 3),
        "sensitivity": _val(cert, 5, 7),
        "sensitivity_unit": _val(cert, 5, 8) or _val(cert, 17, 8),
        "test_level": _val(cert, 6, 7),
        "test_level_unit": _val(cert, 6, 8),
        "customer": _val(cert, 41, 6),
        "as_found": _val(cert, 49, 6),
        "as_left": _val(cert, 50, 6),
        "temperature": _val(cert, 49, 11),
        "humidity": _val(cert, 50, 11),
        "technician": _val(cert, 55, 6),
        "approval": _val(cert, 56, 6),
        "cal_date": _date_val(cert, 55, 11),
        "cal_time": _date_val(cert, 56, 11),
        "due_date": _date_val(cert, 57, 11),
        "iso_method": _val(cert, 32, 5) or "Back-to-Back Comparison per ISO 16063-21",
        "traceability": _val(cert, 30, 5),
        "pvc_info": _val(cert, 31, 5),
        "user_notes": _val(cert, 36, 6),
    }

    # Transducer specs
    info["amp_range"] = _val(cert, 5, 11)
    info["resolution"] = _val(cert, 6, 11)
    info["resonant_freq"] = _val(cert, 7, 11)
    info["temp_range"] = _val(cert, 8, 11)
    info["axis"] = _val(cert, 9, 11)

    # PVC info from raw sheet
    if raw:
        info["pvc_model"] = _val(raw, 3, 4)
        info["pvc_serial"] = _val(raw, 3, 5)
        info["pvc_firmware"] = _val(raw, 3, 6)

    # Reference frequency from data sheet
    if data:
        info["ref_frequency"] = _val(data, 8, 2)

    return info


def read_lin_info(wb):
    """Extract info from Linearity certificate sheets."""
    cert = wb["LINCert"]

    info = {
        "cert_type": "Dynamic Linearity Calibration",
        "model": _val(cert, 5, 5),
        "serial": _val(cert, 6, 5),
        "manufacturer": _val(cert, 7, 5),
        "id_number": _val(cert, 8, 5),
        "description": "",
        "sensitivity": _val(cert, 5, 15),
        "sensitivity_unit": _val(cert, 5, 18),
        "test_frequency": _val(cert, 7, 15),
        "max_linearity": _val(cert, 9, 15),
        "customer": _val(cert, 41, 12),
        "as_found": _val(cert, 48, 12),
        "as_left": _val(cert, 49, 12),
        "temperature": _val(cert, 48, 25),
        "humidity": _val(cert, 50, 25),
        "technician": _val(cert, 55, 15),
        "approval": _val(cert, 56, 15),
        "cal_date": _date_val(cert, 55, 25),
        "due_date": _date_val(cert, 57, 25),
        "iso_method": _val(cert, 33, 12) or "Back-to-Back Comparison per ISO 16063-22",
        "traceability": _val(cert, 30, 12),
        "pvc_info": _val(cert, 31, 12),
        "user_notes": _val(cert, 36, 15),
    }

    # Transducer specs
    info["amp_range"] = _val(cert, 5, 25)
    info["resolution"] = _val(cert, 6, 25)
    info["resonant_freq"] = _val(cert, 7, 25)
    info["temp_range"] = _val(cert, 8, 25)
    info["axis"] = _val(cert, 9, 25)

    if "LINRaw" in wb.sheetnames:
        raw = wb["LINRaw"]
        info["pvc_model"] = _val(raw, 3, 4)
        info["pvc_serial"] = _val(raw, 3, 5)
        info["pvc_firmware"] = _val(raw, 3, 6)

    return info


def read_slin_info(wb):
    """Extract info from Static Linearity certificate sheets."""
    cert = wb["SLINCert"]

    info = {
        "cert_type": "Static Linearity Calibration",
        "model": _val(cert, 5, 5),
        "serial": _val(cert, 6, 5),
        "manufacturer": _val(cert, 7, 5),
        "id_number": _val(cert, 8, 5),
        "description": "",
        "sensitivity": _val(cert, 5, 15),
        "sensitivity_unit": _val(cert, 5, 18),
        "max_linearity": _val(cert, 7, 15),
        "customer": _val(cert, 36, 12),
        "as_found": _val(cert, 47, 12),
        "as_left": _val(cert, 48, 12),
        "temperature": _val(cert, 47, 25),
        "humidity": _val(cert, 48, 25),
        "technician": _val(cert, 52, 15),
        "approval": _val(cert, 53, 15),
        "cal_date": _date_val(cert, 52, 25),
        "due_date": _date_val(cert, 53, 25),
        "iso_method": "Static Linearity Test",
        "traceability": _val(cert, 30, 12),
        "pvc_info": _val(cert, 31, 12),
        "user_notes": _val(cert, 41, 12),
    }

    info["amp_range"] = _val(cert, 5, 25)
    info["resolution"] = _val(cert, 6, 25)
    info["resonant_freq"] = _val(cert, 7, 25)
    info["temp_range"] = _val(cert, 8, 25)

    return info


def read_vib_info(wb):
    """Extract info from Vibration Analyzer certificate sheet."""
    cert = wb["Vib Analyzer Cert"] if "Vib Analyzer Cert" in wb.sheetnames else None
    data = wb["Vib Analyzer Data"] if "Vib Analyzer Data" in wb.sheetnames else None

    # Prefer Data sheet (has actual values); Cert has formulas that may lack cached results
    src = data or cert

    # Lookup tables for coded fields on Data sheet
    MODE_MAP = {"1": "Spectra", "2": "Overall", "3": "Time waveform"}
    FREQ_UNIT_MAP = {"1": "Hz", "2": "CPM"}
    WINDOW_MAP = {"1": "Hanning", "2": "Hamming", "3": "Flattop", "4": "Uniform"}
    SENS_UNIT_MAP = {"1": "mV/EU", "2": "V/EU"}

    if data:
        mode_raw = _val(src, 5, 8)
        analyzer_mode = MODE_MAP.get(mode_raw, mode_raw)
        funit_raw = _val(src, 8, 8)
        freq_unit = FREQ_UNIT_MAP.get(funit_raw, funit_raw)
        win_raw = _val(src, 11, 8)
        window_type = WINDOW_MAP.get(win_raw, win_raw)
    else:
        analyzer_mode = _val(src, 5, 8)
        freq_unit = _val(src, 8, 8)
        window_type = _val(src, 11, 8)

    info = {
        "cert_type": "Vibration Analyzer Certification",
        "manufacturer": _val(src, 5, 3),
        "model": _val(src, 6, 3),
        "serial": _val(src, 7, 3),
        "id_number": "",
        "description": "Vibration Analyzer",
        "cal_tech": _val(src, 8, 3),
        "cal_date": _date_val(src, 9, 3),
        "due_date": _date_val(src, 10, 3),
        "analyzer_mode": analyzer_mode,
        "freq_max": _val(src, 6, 8),
        "freq_min": _val(src, 7, 8),
        "freq_unit": freq_unit,
        "lines_of_resolution": _val(src, 9, 8),
        "averaging_points": _val(src, 10, 8),
        "window_type": window_type,
        "sensor_sensitivity": _val(src, 12, 8),
        "customer": _val(cert or src, 32, 7),
        "iso_method": "Back-to-Back Comparison per ISO 16063-21",
        "traceability": "NIST and PTB",
        "technician": _val(src, 8, 3),
    }

    # Test equipment: PVC info (Data: row 16, Cert: row 15)
    if data:
        info["pvc_model"] = _val(src, 16, 3)
        info["pvc_serial"] = _val(src, 16, 4)
    else:
        info["pvc_model"] = _val(src, 15, 3)
        info["pvc_serial"] = _val(src, 15, 4)

    # Fill defaults
    for key in ["temperature", "humidity", "as_found", "as_left",
                "approval", "user_notes", "pvc_info", "pvc_firmware",
                "sensitivity", "sensitivity_unit", "amp_range",
                "resolution", "resonant_freq", "temp_range", "axis"]:
        info.setdefault(key, "")

    return info


def read_workbook(filepath):
    """Read the filled PVC workbook and return structured info dict."""
    wb = openpyxl.load_workbook(filepath, data_only=True, keep_vba=True)
    cert_type = detect_cert_type(wb)

    readers = {
        "FR": read_fr_info,
        "LIN": read_lin_info,
        "SLIN": read_slin_info,
        "VIB": read_vib_info,
    }

    info = readers[cert_type](wb)
    info["_cert_tag"] = cert_type

    # Fill defaults for missing keys
    defaults = {
        "model": "", "serial": "", "manufacturer": "", "id_number": "",
        "description": "", "customer": "", "technician": "", "approval": "",
        "cal_date": "", "due_date": "", "temperature": "", "humidity": "",
        "as_found": "", "as_left": "", "iso_method": "", "traceability": "",
        "pvc_info": "", "pvc_model": "", "pvc_serial": "", "pvc_firmware": "",
        "user_notes": "", "sensitivity": "", "sensitivity_unit": "",
        "amp_range": "", "resolution": "", "resonant_freq": "", "temp_range": "",
        "axis": "",
    }
    for k, v in defaults.items():
        info.setdefault(k, v)

    return info


# ── PDF Cover Page Builder (matching TEST_Certificate.pdf design) ────────

def make_header_footer(info):
    """Header/footer drawn on every page — same as TEST_Certificate.pdf."""
    def header_footer(canvas_obj, doc):
        canvas_obj.saveState()

        # Logo — top-left
        if os.path.exists(LOGO_PATH):
            try:
                canvas_obj.drawImage(
                    LOGO_PATH,
                    2 * mm, PAGE_H - 22 * mm,
                    width=60 * mm, height=21 * mm,
                    preserveAspectRatio=True, mask='auto'
                )
            except Exception:
                canvas_obj.setFillColor(WEARCHECK_BLACK)
                canvas_obj.setFont("Helvetica-Bold", 14)
                canvas_obj.drawString(MARGIN, PAGE_H - 11 * mm, "WEARCHECK")
        else:
            canvas_obj.setFillColor(WEARCHECK_BLACK)
            canvas_obj.setFont("Helvetica-Bold", 14)
            canvas_obj.drawString(MARGIN, PAGE_H - 11 * mm, "WEARCHECK")

        # Header text — right side
        canvas_obj.setFillColor(WEARCHECK_BLACK)
        canvas_obj.setFont("Helvetica-Bold", 11)
        canvas_obj.drawRightString(PAGE_W - MARGIN, PAGE_H - 8 * mm, "CALIBRATION CERTIFICATE")
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.setFillColor(WEARCHECK_GREY)
        canvas_obj.drawRightString(PAGE_W - MARGIN, PAGE_H - 11.5 * mm,
                                   "WearCheck ARC \u2014 Condition Monitoring Division")

        # Accreditation
        canvas_obj.setFont("Helvetica", 6.5)
        canvas_obj.setFillColor(WEARCHECK_BLACK)
        canvas_obj.drawRightString(PAGE_W - MARGIN, PAGE_H - 14.5 * mm,
                                   "ISO 16063-21  |  NMISA")

        # Footer line
        canvas_obj.setStrokeColor(colors.lightgrey)
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(MARGIN, 14 * mm, PAGE_W - MARGIN, 14 * mm)

        # Footer — equipment details
        canvas_obj.setFillColor(WEARCHECK_BLACK)
        canvas_obj.setFont("Helvetica", 6.5)
        cert_num = f"PVC-{info.get('serial', '')}-{info.get('cal_date', '')}"
        equip_line = (
            f"{info.get('description', '')}  |  "
            f"S/N: {info.get('serial', '')}  |  "
            f"Cert: {cert_num}  |  "
            f"Cal Date: {info.get('cal_date', '')}"
        )
        canvas_obj.drawString(MARGIN, 10 * mm, equip_line)
        canvas_obj.drawRightString(
            PAGE_W - MARGIN, 10 * mm,
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Page {doc.page}"
        )

        canvas_obj.setFont("Helvetica", 6)
        canvas_obj.setFillColor(WEARCHECK_GREY)
        canvas_obj.drawString(
            MARGIN, 6.5 * mm,
            "This certificate is issued in accordance with WearCheck ARC quality management "
            "procedures aligned with ISO/IEC 17025."
        )

        canvas_obj.restoreState()
    return header_footer


def build_section_header(text, styles):
    """Light grey section header bar — same as TEST_Certificate.pdf."""
    return Table(
        [[Paragraph(text, styles["SectionHead"])]],
        colWidths=[PAGE_W - 2 * MARGIN],
        rowHeights=[6 * mm],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), LIGHT_GREY),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ])
    )


def build_sub_section_header(text, styles):
    """Sub-section header — lighter style."""
    return Table(
        [[Paragraph(text, styles["SubSectionHead"])]],
        colWidths=[PAGE_W - 2 * MARGIN],
        rowHeights=[5 * mm],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F5F5")),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ])
    )


def build_field_table(fields, styles):
    """Two-column label/value pairs — same layout as TEST_Certificate.pdf."""
    half = (PAGE_W - 2 * MARGIN) / 2
    col_widths = [half * 0.4, half * 0.6, half * 0.4, half * 0.6]

    rows = []
    for i in range(0, len(fields), 2):
        row = [
            Paragraph(fields[i][0], styles["FieldLabel"]),
            Paragraph(str(fields[i][1]), styles["FieldValue"]),
        ]
        if i + 1 < len(fields):
            row.append(Paragraph(fields[i + 1][0], styles["FieldLabel"]))
            row.append(Paragraph(str(fields[i + 1][1]), styles["FieldValue"]))
        else:
            row.extend(["", ""])
        rows.append(row)

    t = Table(rows, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
    ]))
    return t


def build_signature_table(info, styles):
    """Signature table — matching TEST_Certificate.pdf layout."""
    sig_lbl = ParagraphStyle("sig_lbl", parent=styles["SmallItalic"], alignment=TA_LEFT)
    total_w = PAGE_W - 2 * MARGIN
    gap = 4 * mm
    # Two blocks: Calibrated By + Approved By (no Reviewer for PVC)
    block_w = (total_w - gap) / 2
    name_w = block_w * 0.55
    title_w = block_w * 0.45
    col_widths = [name_w, title_w, gap, name_w, title_w]

    E = ""  # gap column

    sig_data = [
        # Row 0: Role headers
        [Paragraph("<b>Calibrated By:</b>", styles["FieldLabel"]), "", E,
         Paragraph("<b>Approved By:</b>", styles["FieldLabel"]), ""],
        # Row 1: Name
        [Paragraph(info.get("technician", ""), styles["FieldValue"]),
         "", E,
         Paragraph(info.get("approval", ""), styles["FieldValue"]),
         ""],
        # Row 2: Sub-labels
        [Paragraph("Name", sig_lbl), Paragraph("Signature", sig_lbl), E,
         Paragraph("Name", sig_lbl), Paragraph("Signature", sig_lbl)],
        # Row 3: spacer for signature
        ["", "", E, "", ""],
        # Row 4: Date
        [Paragraph(f"Date: {info.get('cal_date', '')}", styles["FieldValue"]),
         "", E,
         Paragraph("Date:", styles["FieldValue"]),
         ""],
        # Row 5: Sub-labels
        [Paragraph("Date Signed", sig_lbl), Paragraph("", sig_lbl), E,
         Paragraph("Date Signed", sig_lbl), Paragraph("", sig_lbl)],
    ]

    t = Table(sig_data, colWidths=col_widths,
              rowHeights=[None, None, None, 10 * mm, None, None])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        # Span role headers
        ("SPAN", (0, 0), (1, 0)),
        ("SPAN", (3, 0), (4, 0)),
        # Lines under name (row 1)
        ("LINEBELOW", (0, 1), (0, 1), 0.5, colors.black),
        ("LINEBELOW", (1, 3), (1, 3), 0.5, colors.black),
        ("LINEBELOW", (3, 1), (3, 1), 0.5, colors.black),
        ("LINEBELOW", (4, 3), (4, 3), 0.5, colors.black),
    ]))
    return t


def generate_cover_page(info, output_path):
    """Build the single-page cover page PDF — same design as TEST_Certificate.pdf."""
    styles = get_styles()

    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        topMargin=20 * mm,
        bottomMargin=15 * mm,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        title="Calibration Certificate Cover Page",
        author="WearCheck ARC",
    )

    elements = []

    # Certificate title
    elements.append(Spacer(1, 1 * mm))
    elements.append(Paragraph("CERTIFICATE OF CALIBRATION", styles["CertTitle"]))
    elements.append(HRFlowable(
        width="100%", thickness=0.5, color=colors.lightgrey, spaceAfter=1.5 * mm
    ))

    # ── Certificate Information ──
    cert_num = f"PVC-{info['serial']}-{info['cal_date']}" if info["serial"] else "\u2014"
    elements.append(build_section_header("CERTIFICATE INFORMATION", styles))
    elements.append(build_field_table([
        ("Certificate No:", cert_num),
        ("Date of Issue:", datetime.now().strftime("%Y-%m-%d")),
        ("Calibration Date:", info.get("cal_date", "")),
        ("Next Due:", info.get("due_date", "")),
        ("Customer:", info.get("customer", "")),
        ("", ""),
    ], styles))

    # ── Sensor Information ──
    elements.append(build_section_header("SENSOR INFORMATION", styles))
    sensor_fields = [
        ("Description:", info.get("description", "")),
        ("Manufacturer:", info.get("manufacturer", "")),
        ("Model:", info.get("model", "")),
        ("Serial No:", info.get("serial", "")),
    ]
    if info.get("id_number"):
        sensor_fields += [("ID Number:", info["id_number"]), ("", "")]
    elements.append(build_field_table(sensor_fields, styles))

    # ── Transducer Specifications ──
    spec_fields = [
        ("Amp. Range:", info.get("amp_range", "")),
        ("Resolution:", info.get("resolution", "")),
        ("Resonant Freq:", info.get("resonant_freq", "")),
        ("Temp. Range:", info.get("temp_range", "")),
    ]
    if any(v for _, v in spec_fields):
        elements.append(build_section_header("TRANSDUCER SPECIFICATIONS", styles))
        elements.append(build_field_table(spec_fields, styles))

    # ── Calibration Data Summary ──
    elements.append(build_section_header("CALIBRATION DATA", styles))
    cal_summary = []
    if info.get("sensitivity"):
        sens_str = info["sensitivity"]
        if info.get("sensitivity_unit"):
            sens_str = f"{sens_str} {info['sensitivity_unit']}"
        cal_summary += [("Sensitivity:", sens_str)]
    else:
        cal_summary += [("Sensitivity:", "")]

    if info.get("ref_frequency"):
        cal_summary += [("Ref. Frequency:", f"{info['ref_frequency']} Hz")]
    else:
        cal_summary += [("Ref. Frequency:", "")]

    if info.get("test_level"):
        unit = info.get("test_level_unit", "")
        cal_summary += [("Test Level:", f"{info['test_level']} {unit}".strip()), ("", "")]
    if info.get("test_frequency"):
        cal_summary += [("Test Frequency:", info["test_frequency"]), ("", "")]
    if info.get("max_linearity"):
        cal_summary += [("Max Linearity:", f"{info['max_linearity']}%"), ("", "")]
    if info.get("analyzer_mode"):
        cal_summary += [("Analyzer Mode:", info["analyzer_mode"]), ("", "")]

    if len(cal_summary) % 2 != 0:
        cal_summary.append(("", ""))
    elements.append(build_field_table(cal_summary, styles))

    # ── Conditions ──
    elements.append(build_section_header("CONDITIONS", styles))
    cond_fields = []
    if info.get("as_found"):
        cond_fields += [("As Found:", info["as_found"])]
    if info.get("as_left"):
        cond_fields += [("As Left:", info["as_left"])]
    if info.get("temperature"):
        cond_fields += [("Temperature:", f"{info['temperature']} \u00b0C")]
    if info.get("humidity"):
        cond_fields += [("Humidity:", f"{info['humidity']} %")]
    if not cond_fields:
        cond_fields = [("Temperature:", ""), ("Humidity:", "")]
    if len(cond_fields) % 2 != 0:
        cond_fields.append(("", ""))
    elements.append(build_field_table(cond_fields, styles))

    # ── Procedure ──
    elements.append(build_section_header("PROCEDURE", styles))
    proc_text = info.get("pvc_info", "")
    iso = info.get("iso_method", "") or "Back-to-Back Comparison per ISO 16063-21"
    if proc_text:
        proc_text = f"{proc_text}<br/>{iso}"
    else:
        proc_text = iso
    elements.append(Paragraph(proc_text, styles["FieldValue"]))

    # Reference equipment
    ref_parts = ["Portable Vibration Calibrator (PVC)"]
    if info.get("pvc_model"):
        ref_parts.append(f"M/N: {info['pvc_model']}")
    if info.get("pvc_serial"):
        ref_parts.append(f"S/N: {info['pvc_serial']}")
    if info.get("pvc_firmware"):
        ref_parts.append(f"FW: {info['pvc_firmware']}")
    elements.append(build_sub_section_header("Reference Equipment", styles))
    elements.append(build_field_table([
        ("Instrument:", ", ".join(ref_parts)),
        ("", ""),
    ], styles))

    # Traceability & uncertainty
    elements.append(Paragraph(
        "<b>Traceability:</b> The measurements reported herein are traceable to NIST (USA) "
        "and PTB (Germany) through an unbroken chain of calibrations.",
        styles["SmallItalic"]
    ))
    elements.append(Paragraph(
        "<b>Uncertainty:</b> The reported expanded uncertainty is based on a standard uncertainty "
        "multiplied by a coverage factor k=2, providing a level of confidence of approximately 95%.",
        styles["SmallItalic"]
    ))

    if info.get("user_notes"):
        elements.append(Spacer(1, 1 * mm))
        elements.append(Paragraph(f"<b>Notes:</b> {info['user_notes']}", styles["SmallItalic"]))

    elements.append(Spacer(1, 1 * mm))

    # ── Authorisation ──
    elements.append(build_section_header("AUTHORISATION", styles))
    elements.append(build_signature_table(info, styles))

    # Build with header/footer
    hf = make_header_footer(info)
    doc.build(elements, onFirstPage=hf, onLaterPages=hf)
    return output_path


# ── Main ──────────────────────────────────────────────────────────────────

def main():
    if len(sys.argv) < 2:
        print("Usage: python generate_cover_page.py <filled_workbook.xlsm>")
        print()
        print("The workbook must be a saved copy of the PVC Report Generation")
        print("Workbook (.xltm) with calibration data filled in.")
        print()
        print("Tip: Open the .xltm in Excel, run the calibration, save as .xlsm,")
        print("     then run this script on the saved file.")
        sys.exit(1)

    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"ERROR: File not found: {filepath}")
        sys.exit(1)

    print(f"Reading workbook: {filepath}")
    info = read_workbook(filepath)

    # Build output filename
    model = info.get("model", "").replace(" ", "_").replace("/", "-") or "Unknown"
    serial = info.get("serial", "").replace(" ", "_") or "Unknown"
    cal_date = info.get("cal_date", datetime.now().strftime("%Y-%m-%d"))
    output_name = f"CoverPage_{model}_{serial}_{cal_date}.pdf"
    output_path = os.path.join(os.path.dirname(os.path.abspath(filepath)), output_name)

    print(f"Certificate type: {info.get('cert_type', 'Unknown')}")
    print(f"Sensor: {info.get('model', '?')} / S/N: {info.get('serial', '?')}")
    print(f"Generating cover page...")

    generate_cover_page(info, output_path)
    print(f"Cover page saved: {output_path}")


if __name__ == "__main__":
    main()
