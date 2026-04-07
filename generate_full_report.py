"""
WearCheck ARC — Full Calibration Report Generator
===================================================
Reads two Vibration Analyzer Excel workbooks (ACC and VEL tests for the
same instrument) and produces a single PDF report:

  Page 1 : Branded Certificate of Calibration cover page
            (populated from both Excel files)
  Page 2+: Original machine-generated PDF reports appended as-is

Usage:
    python generate_full_report.py <data_folder>

    <data_folder> must contain exactly two .xlsx files whose names
    contain "Acc" and "Vel" respectively, and optionally matching .pdf
    files.  If .pdf files are missing the Excel sheets are rendered
    to PDF automatically.

Output:
    <data_folder>/FullReport_<Model>_<Serial>_<Date>.pdf
"""

import os
import sys
import glob
import io
from datetime import datetime, date, time as dt_time

import qrcode

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image,
    HRFlowable, PageBreak,
)
from reportlab.platypus.flowables import KeepTogether

import openpyxl
from PyPDF2 import PdfReader, PdfWriter

# ── Brand Colors ─────────────────────────────────────────────────────────

WEARCHECK_RED   = colors.HexColor("#C2040B")
WEARCHECK_BLACK = colors.HexColor("#1A1A1A")
WEARCHECK_GREY  = colors.HexColor("#4D4D4D")
WEARCHECK_LTGREY = colors.HexColor("#E6E6E6")
LIGHT_GREY      = colors.HexColor("#F2F2F2")
PASS_GREEN      = colors.HexColor("#C6EFCE")
DARK_GREEN      = colors.HexColor("#006100")

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
LOGO_PATH  = os.path.join(SCRIPT_DIR, "WearCheck Logo.png")
PAGE_W, PAGE_H = A4
MARGIN = 15 * mm

NMISA_QR_URL = "https://nmisa.microsoftcrmportals.com/QRCertificates/?data=NMISA-AUV-2026-11287"

_BLANK = {None, "", "#N/A", "#VALUE!", "#REF!", "#DIV/0!", "0", "0.0"}


# ── Helpers ──────────────────────────────────────────────────────────────

def _val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    s = str(v).strip()
    if s in _BLANK:
        return ""
    return s


def _date_val(ws, row, col):
    v = ws.cell(row=row, column=col).value
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, dt_time):
        if v == dt_time(0, 0):
            return ""
        return v.strftime("%H:%M")
    if v is None:
        return ""
    return str(v).strip()


def _num(ws, row, col, decimals=2):
    """Return a numeric cell as a formatted string, or empty."""
    v = ws.cell(row=row, column=col).value
    if v is None:
        return ""
    try:
        return f"{float(v):.{decimals}f}"
    except (ValueError, TypeError):
        return str(v).strip()


# ── Read Vib Analyzer Info from one workbook ────────────────────────────

def read_vib_analyzer(filepath):
    """Return dict with all relevant Vib Analyzer Cert/Data fields."""
    wb = openpyxl.load_workbook(filepath, data_only=True, keep_vba=True)
    cert = wb["Vib Analyzer Cert"] if "Vib Analyzer Cert" in wb.sheetnames else None
    data = wb["Vib Analyzer Data"] if "Vib Analyzer Data" in wb.sheetnames else None

    if not data and not cert:
        wb.close()
        raise ValueError("Neither 'Vib Analyzer Data' nor 'Vib Analyzer Cert' found")

    # Prefer Data sheet (has actual values); Cert has formulas that may lack cached results
    src = data or cert

    # Lookup tables for coded fields on the Data sheet
    MODE_MAP = {"1": "Spectra", "2": "Overall", "3": "Time waveform"}
    FREQ_UNIT_MAP = {"1": "Hz", "2": "CPM"}
    WINDOW_MAP = {"1": "Hanning", "2": "Hamming", "3": "Flattop", "4": "Uniform"}
    SENS_UNIT_MAP = {"1": "mV/EU", "2": "V/EU"}

    info = {}

    # ── Analyzer / Meter Information (same positions on both sheets) ──
    info["manufacturer"] = _val(src, 5, 3)   # C5
    info["model"]        = _val(src, 6, 3)   # C6
    info["serial"]       = _val(src, 7, 3)   # C7
    info["cal_tech"]     = _val(src, 8, 3)   # C8
    info["cal_date"]     = _val(src, 9, 3)   # C9
    info["cal_due"]      = _val(src, 10, 3)  # C10

    # ── Analyzer Settings ──
    if data:
        # Data sheet stores numeric codes for some fields; resolve them
        mode_raw = _val(src, 5, 8)
        info["analyzer_mode"] = MODE_MAP.get(mode_raw, mode_raw)
        info["freq_max"]      = _val(src, 6, 8)
        info["freq_min"]      = _val(src, 7, 8)
        funit_raw = _val(src, 8, 8)
        info["freq_unit"]     = FREQ_UNIT_MAP.get(funit_raw, funit_raw)
        info["lines_resolution"] = _val(src, 9, 8)
        info["avg_points"]    = _val(src, 10, 8)
        win_raw = _val(src, 11, 8)
        info["window_type"]   = WINDOW_MAP.get(win_raw, win_raw)
        info["sensor_input_sens"] = _val(src, 12, 8)
        sunit_raw = _val(src, 13, 8)
        info["sensor_input_unit"] = SENS_UNIT_MAP.get(sunit_raw, sunit_raw)
    else:
        info["analyzer_mode"]  = _val(src, 5, 8)
        info["freq_max"]       = _val(src, 6, 8)
        info["freq_min"]       = _val(src, 7, 8)
        info["freq_unit"]      = _val(src, 8, 8)
        info["lines_resolution"] = _val(src, 9, 8)
        info["avg_points"]     = _val(src, 10, 8)
        info["window_type"]    = _val(src, 11, 8)
        info["sensor_input_sens"] = _val(src, 12, 8)
        info["sensor_input_unit"] = _val(src, 12, 9)

    # ── Test type detection (J4 header — on Cert sheet) ──
    test_src = cert or src
    test_header = _val(test_src, 4, 10)  # J4
    info["test_type"] = test_header

    # ── Test parameters (column L-M, rows 5-8 — on Cert sheet) ──
    param_src = cert or src
    for r in range(5, 9):
        label = _val(param_src, r, 12)
        value = _val(param_src, r, 13)
        if label and value:
            info[f"test_param_{r}"] = (label, value)

    # ── Test Equipment ──
    if data:
        # Data sheet: PVC at row 16, Sensor at row 17
        # Sensitivity at col G(7), unit at H(8), tolerance at K(11)
        info["pvc_model"]       = _val(src, 16, 3)
        info["pvc_serial"]      = _val(src, 16, 4)
        info["pvc_cal_date"]    = _date_val(src, 16, 6)
        info["pvc_sensitivity"] = _num(src, 16, 7)
        info["pvc_sens_unit"]   = _val(src, 16, 8)
        info["pvc_tolerance"]   = _num(src, 16, 11)
        info["pvc_deviation"]   = _num(src, 16, 12)

        info["sensor_model"]       = _val(src, 17, 3)
        info["sensor_serial"]      = _val(src, 17, 4)
        info["sensor_cal_date"]    = _date_val(src, 17, 6)
        info["sensor_sensitivity"] = _num(src, 17, 7)
        info["sensor_sens_unit"]   = _val(src, 17, 8)
        info["sensor_tolerance"]   = _num(src, 17, 11)
        info["sensor_deviation"]   = _num(src, 17, 12)
    else:
        info["pvc_model"]      = _val(src, 15, 3)
        info["pvc_serial"]     = _val(src, 15, 4)
        info["pvc_cal_date"]   = _date_val(src, 15, 6)
        info["pvc_sensitivity"] = _num(src, 15, 8)
        info["pvc_sens_unit"]  = _val(src, 15, 9)
        info["pvc_tolerance"]  = _num(src, 15, 10)
        info["pvc_deviation"]  = _num(src, 15, 12)

        info["sensor_model"]     = _val(src, 16, 3)
        info["sensor_serial"]    = _val(src, 16, 4)
        info["sensor_cal_date"]  = _date_val(src, 16, 6)
        info["sensor_sensitivity"] = _num(src, 16, 8)
        info["sensor_sens_unit"] = _val(src, 16, 9)
        info["sensor_tolerance"] = _num(src, 16, 10)
        info["sensor_deviation"] = _num(src, 16, 12)

    # ── Customer (literal value on Cert sheet) ──
    info["customer"] = _val(cert or src, 33, 7)  # G33

    # ── Data Table ──
    data_src = cert or src
    data_rows = []
    for r in range(34, data_src.max_row + 1):
        b = data_src.cell(row=r, column=2).value
        c = data_src.cell(row=r, column=3).value
        d = data_src.cell(row=r, column=4).value
        if b is not None and c is not None:
            data_rows.append((b, c, d))
    info["data_table"] = data_rows

    # Data table column headers
    info["data_col1"] = _val(data_src, 33, 2)  # B33
    info["data_col2"] = _val(data_src, 33, 3)  # C33
    info["data_col3"] = _val(data_src, 33, 4)  # D33

    # ── Note (literal value on Cert sheet) ──
    info["note"] = _val(cert or src, 41, 7)  # G41

    # ── Max deviation (for Linearity test) ──
    info["abs_max_deviation"] = _val(param_src, 8, 13)  # M8 only for linearity

    wb.close()
    return info


# ── Merge info from both workbooks ──────────────────────────────────────

def merge_info(acc_info=None, vel_info=None):
    """Combine ACC and/or VEL info into a single dict for the cover page."""
    merged = {}

    sources = [s for s in [acc_info, vel_info] if s is not None]
    if not sources:
        return merged

    # Common analyzer info (take from whichever is available)
    for key in ["manufacturer", "model", "serial", "cal_tech", "cal_date",
                "cal_due", "analyzer_mode", "freq_max", "freq_min",
                "freq_unit", "lines_resolution", "avg_points", "window_type",
                "sensor_input_sens", "sensor_input_unit", "customer",
                "pvc_model", "pvc_serial", "pvc_cal_date", "pvc_sensitivity",
                "pvc_sens_unit", "pvc_tolerance", "pvc_deviation",
                "sensor_model", "sensor_serial", "sensor_cal_date",
                "sensor_sensitivity", "sensor_sens_unit", "sensor_tolerance",
                "sensor_deviation", "note"]:
        for src in sources:
            val = src.get(key)
            if val:
                merged[key] = val
                break
        else:
            merged[key] = ""

    # Test-specific info kept separate
    if acc_info:
        merged["acc_test_type"] = acc_info.get("test_type", "Linearity Test")
        merged["acc_params"] = {}
        for r in range(5, 9):
            k = f"test_param_{r}"
            if k in acc_info:
                merged["acc_params"][acc_info[k][0]] = acc_info[k][1]
    else:
        merged["acc_test_type"] = ""
        merged["acc_params"] = {}

    if vel_info:
        merged["vel_test_type"] = vel_info.get("test_type", "Frequency Response Test")
        merged["vel_params"] = {}
        for r in range(5, 9):
            k = f"test_param_{r}"
            if k in vel_info:
                merged["vel_params"][vel_info[k][0]] = vel_info[k][1]
    else:
        merged["vel_test_type"] = ""
        merged["vel_params"] = {}

    return merged


# ── Styles ───────────────────────────────────────────────────────────────

def get_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="CertTitle", fontName="Helvetica-Bold", fontSize=16,
        textColor=WEARCHECK_BLACK, alignment=TA_CENTER, spaceAfter=2 * mm,
    ))
    styles.add(ParagraphStyle(
        name="SectionHead", fontName="Helvetica-Bold", fontSize=9,
        textColor=WEARCHECK_BLACK, alignment=TA_LEFT,
        leftIndent=2 * mm, spaceBefore=1.5 * mm, spaceAfter=0.5 * mm,
    ))
    styles.add(ParagraphStyle(
        name="SubSectionHead", fontName="Helvetica-Bold", fontSize=8,
        textColor=WEARCHECK_GREY, alignment=TA_LEFT,
        leftIndent=4 * mm, spaceBefore=1 * mm, spaceAfter=0.5 * mm,
    ))
    styles.add(ParagraphStyle(
        name="FieldLabel", fontName="Helvetica-Bold", fontSize=7.5,
        textColor=WEARCHECK_BLACK, alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        name="FieldValue", fontName="Helvetica", fontSize=8,
        alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        name="SmallItalic", fontName="Helvetica-Oblique", fontSize=6.5,
        textColor=colors.grey, alignment=TA_JUSTIFY, spaceBefore=1 * mm,
    ))
    styles.add(ParagraphStyle(
        name="Footer", fontName="Helvetica", fontSize=6,
        textColor=colors.grey, alignment=TA_CENTER,
    ))
    return styles


# ── Reusable PDF building blocks ────────────────────────────────────────

def build_section_header(text, styles):
    return Table(
        [[Paragraph(text, styles["SectionHead"])]],
        colWidths=[PAGE_W - 2 * MARGIN],
        rowHeights=[6 * mm],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), LIGHT_GREY),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ])
    )


def build_sub_section_header(text, styles):
    return Table(
        [[Paragraph(text, styles["SubSectionHead"])]],
        colWidths=[PAGE_W - 2 * MARGIN],
        rowHeights=[5 * mm],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F5F5")),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ])
    )


def build_field_table(fields, styles):
    half = (PAGE_W - 2 * MARGIN) / 2
    col_widths = [half * 0.4, half * 0.6, half * 0.4, half * 0.6]

    rows = []
    for i in range(0, len(fields), 2):
        row = [
            Paragraph(fields[i][0], styles["FieldLabel"]),
            Paragraph(str(fields[i][1]), styles["FieldValue"]),
        ]
        if i + 1 < len(fields):
            row.append(Paragraph(fields[i + 1][0], styles["FieldLabel"]))
            row.append(Paragraph(str(fields[i + 1][1]), styles["FieldValue"]))
        else:
            row.extend(["", ""])
        rows.append(row)

    t = Table(rows, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
    ]))
    return t


def build_field_table_stacked(fields, styles):
    """Full-width rows (one field per row) to avoid overlap with long values."""
    total_w = PAGE_W - 2 * MARGIN
    col_widths = [total_w * 0.18, total_w * 0.82]

    rows = []
    for label, value in fields:
        rows.append([
            Paragraph(label, styles["FieldLabel"]),
            Paragraph(str(value), styles["FieldValue"]),
        ])

    t = Table(rows, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
    ]))
    return t


def build_signature_table(info, styles):
    sig_lbl = ParagraphStyle("sig_lbl", parent=styles["SmallItalic"], alignment=TA_LEFT)
    total_w = PAGE_W - 2 * MARGIN
    gap = 4 * mm
    block_w = (total_w - gap) / 2
    name_w = block_w * 0.55
    title_w = block_w * 0.45
    col_widths = [name_w, title_w, gap, name_w, title_w]
    E = ""

    sig_data = [
        [Paragraph("<b>Calibrated By:</b>", styles["FieldLabel"]), "", E,
         Paragraph("<b>Approved By:</b>", styles["FieldLabel"]), ""],
        [Paragraph(info.get("cal_tech", ""), styles["FieldValue"]),
         "", E, Paragraph("", styles["FieldValue"]), ""],
        [Paragraph("Name", sig_lbl), Paragraph("Signature", sig_lbl), E,
         Paragraph("Name", sig_lbl), Paragraph("Signature", sig_lbl)],
        ["", "", E, "", ""],
        [Paragraph(f"Date: {datetime.now().strftime('%d/%m/%Y')}", styles["FieldValue"]),
         "", E, Paragraph("Date:", styles["FieldValue"]), ""],
        [Paragraph("Date Signed", sig_lbl), Paragraph("", sig_lbl), E,
         Paragraph("Date Signed", sig_lbl), Paragraph("", sig_lbl)],
    ]

    t = Table(sig_data, colWidths=col_widths,
              rowHeights=[None, None, None, 10 * mm, None, None])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ("SPAN", (0, 0), (1, 0)),
        ("SPAN", (3, 0), (4, 0)),
        # Name lines — span both columns per block
        ("LINEBELOW", (0, 1), (1, 1), 0.5, colors.black),
        ("LINEBELOW", (3, 1), (4, 1), 0.5, colors.black),
        # Signature lines — span both columns per block
        ("LINEBELOW", (0, 3), (1, 3), 0.5, colors.black),
        ("LINEBELOW", (3, 3), (4, 3), 0.5, colors.black),
    ]))
    return t


# ── Header / Footer ─────────────────────────────────────────────────────

def make_header_footer(info):
    def draw(canvas_obj, doc):
        canvas_obj.saveState()

        # Logo top-left
        if os.path.exists(LOGO_PATH):
            try:
                logo_h = 20 * mm
                logo_w = logo_h * 488 / 511
                canvas_obj.drawImage(
                    LOGO_PATH, MARGIN, PAGE_H - MARGIN - logo_h,
                    width=logo_w, height=logo_h,
                    preserveAspectRatio=True, mask='auto',
                )
            except Exception:
                canvas_obj.setFillColor(WEARCHECK_BLACK)
                canvas_obj.setFont("Helvetica-Bold", 14)
                canvas_obj.drawString(MARGIN, PAGE_H - 11 * mm, "WEARCHECK")
        else:
            canvas_obj.setFillColor(WEARCHECK_BLACK)
            canvas_obj.setFont("Helvetica-Bold", 14)
            canvas_obj.drawString(MARGIN, PAGE_H - 11 * mm, "WEARCHECK")

        # Header right
        canvas_obj.setFillColor(WEARCHECK_BLACK)
        canvas_obj.setFont("Helvetica-Bold", 11)
        canvas_obj.drawRightString(PAGE_W - MARGIN, PAGE_H - 8 * mm,
                                   "CALIBRATION CERTIFICATE")
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.setFillColor(WEARCHECK_GREY)
        canvas_obj.drawRightString(PAGE_W - MARGIN, PAGE_H - 11.5 * mm,
                                   "WearCheck ARC \u2014 Condition Monitoring Division")
        canvas_obj.setFont("Helvetica", 6.5)
        canvas_obj.setFillColor(WEARCHECK_BLACK)
        canvas_obj.drawRightString(PAGE_W - MARGIN, PAGE_H - 14.5 * mm,
                                   "ISO 16063-21  |  NMISA-AUV-2026-11287")

        # QR code — between logo and header text
        try:
            qr = qrcode.make(NMISA_QR_URL, box_size=10, border=0)
            qr_buf = io.BytesIO()
            qr.save(qr_buf, format='PNG')
            qr_buf.seek(0)
            qr_size = 14 * mm
            qr_x = MARGIN + 22 * mm
            qr_y = PAGE_H - MARGIN - qr_size
            from reportlab.lib.utils import ImageReader
            canvas_obj.drawImage(ImageReader(qr_buf), qr_x, qr_y,
                                 width=qr_size, height=qr_size)
            canvas_obj.setFont("Helvetica", 4.5)
            canvas_obj.setFillColor(WEARCHECK_GREY)
            canvas_obj.drawCentredString(qr_x + qr_size / 2, qr_y - 3 * mm,
                                         "Scan to authenticate")
        except Exception:
            pass

        # Footer line
        canvas_obj.setStrokeColor(colors.lightgrey)
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(MARGIN, 14 * mm, PAGE_W - MARGIN, 14 * mm)

        serial = info.get("serial", "")
        cal_date = info.get("cal_date", "")
        cert_num = f"VIB-{serial}-{cal_date}" if serial else "\u2014"
        model = info.get("model", "")
        mfr = info.get("manufacturer", "")
        equip_line = (
            f"{mfr} {model}  |  S/N: {serial}  |  "
            f"Cert: {cert_num}  |  Cal Date: {cal_date}"
        )
        canvas_obj.setFillColor(WEARCHECK_BLACK)
        canvas_obj.setFont("Helvetica", 6.5)
        canvas_obj.drawString(MARGIN, 10 * mm, equip_line)
        canvas_obj.drawRightString(
            PAGE_W - MARGIN, 10 * mm,
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Page {doc.page}",
        )
        canvas_obj.setFont("Helvetica", 6)
        canvas_obj.setFillColor(WEARCHECK_GREY)
        canvas_obj.drawString(
            MARGIN, 6.5 * mm,
            "This certificate is issued in accordance with WearCheck ARC quality "
            "management procedures aligned with ISO/IEC 17025.",
        )
        canvas_obj.restoreState()
    return draw


# ── Cover Page Builder ──────────────────────────────────────────────────

def generate_cover_page(info, output_path):
    """Build a single-page cover page PDF for both ACC + VEL tests."""
    styles = get_styles()

    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        topMargin=20 * mm, bottomMargin=15 * mm,
        leftMargin=MARGIN, rightMargin=MARGIN,
        title="Calibration Certificate — Vibration Analyzer",
        author="WearCheck ARC",
    )

    elements = []

    # Title
    elements.append(Spacer(1, 1 * mm))
    elements.append(Paragraph("CERTIFICATE OF CALIBRATION", styles["CertTitle"]))
    elements.append(HRFlowable(
        width="100%", thickness=0.5, color=colors.lightgrey, spaceAfter=1.5 * mm,
    ))

    # ── Certificate Information ──
    serial = info.get("serial", "")
    cal_date = info.get("cal_date", "")
    cert_num = f"VIB-{serial}-{cal_date}" if serial else "\u2014"
    elements.append(build_section_header("CERTIFICATE INFORMATION", styles))
    elements.append(build_field_table([
        ("Certificate No:", cert_num),
        ("Date of Issue:", datetime.now().strftime("%Y-%m-%d")),
        ("Calibration Date:", cal_date),
        ("Next Due:", info.get("cal_due", "")),
        ("Customer:", info.get("customer", "")),
        ("Calibration Technician:", info.get("cal_tech", "")),
    ], styles))

    # ── Analyzer Information ──
    elements.append(build_section_header("ANALYZER / METER INFORMATION", styles))
    elements.append(build_field_table([
        ("Manufacturer:", info.get("manufacturer", "")),
        ("Model:", info.get("model", "")),
        ("Serial No:", serial),
        ("Description:", "Vibration Analyzer"),
    ], styles))

    # ── Analyzer Settings ──
    elements.append(build_section_header("ANALYZER SETTINGS", styles))
    elements.append(build_field_table([
        ("Analyzer Mode:", info.get("analyzer_mode", "")),
        ("Frequency Max:", f"{info.get('freq_max', '')} {info.get('freq_unit', '')}"),
        ("Frequency Min:", f"{info.get('freq_min', '')} {info.get('freq_unit', '')}"),
        ("Lines of Resolution:", info.get("lines_resolution", "")),
        ("Averaging Points:", info.get("avg_points", "")),
        ("Window Type:", info.get("window_type", "")),
        ("Sensor Input Sensitivity:",
         f"{info.get('sensor_input_sens', '')} {info.get('sensor_input_unit', '')}"),
        ("", ""),
    ], styles))

    # ── Tests Performed ──
    elements.append(build_section_header("TESTS PERFORMED", styles))

    test_num = 0
    # ACC (Linearity) test
    acc_type = info.get("acc_test_type", "")
    if acc_type:
        test_num += 1
        elements.append(build_sub_section_header(f"Test {test_num}: {acc_type}", styles))
        acc_fields = []
        for label, value in info.get("acc_params", {}).items():
            acc_fields.append((f"{label}", str(value)))
        if len(acc_fields) % 2 != 0:
            acc_fields.append(("", ""))
        if acc_fields:
            elements.append(build_field_table(acc_fields, styles))

    # VEL (Frequency Response) test
    vel_type = info.get("vel_test_type", "")
    if vel_type:
        test_num += 1
        elements.append(build_sub_section_header(f"Test {test_num}: {vel_type}", styles))
        vel_fields = []
        for label, value in info.get("vel_params", {}).items():
            vel_fields.append((f"{label}", str(value)))
        if len(vel_fields) % 2 != 0:
            vel_fields.append(("", ""))
        if vel_fields:
            elements.append(build_field_table(vel_fields, styles))

    # ── Test Equipment ──
    elements.append(build_section_header("TEST EQUIPMENT", styles))

    pvc_line = f"PVC M/N: {info.get('pvc_model', '')}, S/N: {info.get('pvc_serial', '')}"
    if info.get("pvc_cal_date"):
        pvc_line += f", Cal: {info['pvc_cal_date']}"
    sensor_line = (
        f"Sensor M/N: {info.get('sensor_model', '')}, "
        f"S/N: {info.get('sensor_serial', '')}"
    )
    if info.get("sensor_cal_date"):
        sensor_line += f", Cal: {info['sensor_cal_date']}"

    elements.append(build_field_table_stacked([
        ("PVC:", pvc_line),
        ("Sensitivity:", f"{info.get('pvc_sensitivity', '')} {info.get('pvc_sens_unit', '')}"),
        ("Sensor:", sensor_line),
        ("Sensitivity:", f"{info.get('sensor_sensitivity', '')} {info.get('sensor_sens_unit', '')}"),
    ], styles))

    # ── Procedure ──
    elements.append(build_section_header("PROCEDURE", styles))
    elements.append(Paragraph(
        "Back-to-Back Comparison per ISO 16063-21",
        styles["FieldValue"],
    ))
    elements.append(Paragraph(
        "<b>Traceability:</b> The measurements reported herein are traceable to "
        "NIST (USA) and PTB (Germany) through an unbroken chain of calibrations.",
        styles["SmallItalic"],
    ))
    elements.append(Paragraph(
        "<b>Uncertainty:</b> The reported expanded uncertainty is based on a "
        "standard uncertainty multiplied by a coverage factor k=2, providing a "
        "level of confidence of approximately 95%.",
        styles["SmallItalic"],
    ))

    if info.get("note"):
        elements.append(Spacer(1, 1 * mm))
        elements.append(Paragraph(
            f"<b>Note:</b> {info['note']}", styles["SmallItalic"],
        ))

    elements.append(Spacer(1, 1 * mm))

    # ── Authorisation ──
    elements.append(build_section_header("AUTHORISATION", styles))
    elements.append(build_signature_table(info, styles))

    # Build
    hf = make_header_footer(info)
    doc.build(elements, onFirstPage=hf, onLaterPages=hf)
    return output_path


# ── Merge PDFs ──────────────────────────────────────────────────────────

def generate_data_page(info, output_path):
    """Render the Vib Analyzer Cert data table as a branded PDF page.

    Used as a fallback when no machine-generated PDF is available.
    """
    styles = get_styles()
    styles.add(ParagraphStyle(
        name="DataTitle", fontName="Helvetica-Bold", fontSize=12,
        textColor=WEARCHECK_BLACK, alignment=TA_CENTER, spaceAfter=3 * mm,
    ))
    styles.add(ParagraphStyle(
        name="CellText", fontName="Helvetica", fontSize=7.5,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="CellHead", fontName="Helvetica-Bold", fontSize=7.5,
        textColor=colors.white, alignment=TA_CENTER,
    ))

    doc = SimpleDocTemplate(
        output_path, pagesize=A4,
        topMargin=20 * mm, bottomMargin=15 * mm,
        leftMargin=MARGIN, rightMargin=MARGIN,
    )

    elements = []
    elements.append(Spacer(1, 2 * mm))

    test_type = info.get("test_type", "Vibration Analyzer Certification")
    elements.append(Paragraph(
        f"~ {test_type} ~", styles["DataTitle"],
    ))

    # ── Analyzer Info block ──
    elements.append(build_section_header("ANALYZER / METER INFORMATION", styles))
    elements.append(build_field_table([
        ("Manufacturer:", info.get("manufacturer", "")),
        ("Model:", info.get("model", "")),
        ("Serial No:", info.get("serial", "")),
        ("Cal Tech:", info.get("cal_tech", "")),
        ("Cal Date:", info.get("cal_date", "")),
        ("Cal Due:", info.get("cal_due", "")),
    ], styles))

    # ── Settings ──
    elements.append(build_section_header("ANALYZER SETTINGS", styles))
    elements.append(build_field_table([
        ("Analyzer Mode:", info.get("analyzer_mode", "")),
        ("Frequency Max:", f"{info.get('freq_max', '')} {info.get('freq_unit', '')}"),
        ("Frequency Min:", f"{info.get('freq_min', '')} {info.get('freq_unit', '')}"),
        ("Lines of Resolution:", info.get("lines_resolution", "")),
        ("Averaging Points:", info.get("avg_points", "")),
        ("Window Type:", info.get("window_type", "")),
        ("Sensor Input Sensitivity:",
         f"{info.get('sensor_input_sens', '')} {info.get('sensor_input_unit', '')}"),
        ("", ""),
    ], styles))

    # ── Test Parameters ──
    elements.append(build_section_header("TEST PARAMETERS", styles))
    param_fields = []
    for r in range(5, 9):
        k = f"test_param_{r}"
        if k in info:
            param_fields.append((info[k][0], str(info[k][1])))
    if param_fields:
        if len(param_fields) % 2 != 0:
            param_fields.append(("", ""))
        elements.append(build_field_table(param_fields, styles))

    # ── Test Equipment ──
    elements.append(build_section_header("TEST EQUIPMENT", styles))
    equip_total_w = PAGE_W - 2 * MARGIN
    equip_cols = [equip_total_w * 0.14, equip_total_w * 0.12, equip_total_w * 0.12,
                  equip_total_w * 0.16, equip_total_w * 0.18, equip_total_w * 0.14,
                  equip_total_w * 0.14]
    equip_header = [
        Paragraph("Description", styles["CellHead"]),
        Paragraph("M/N", styles["CellHead"]),
        Paragraph("S/N", styles["CellHead"]),
        Paragraph("Cal. Date", styles["CellHead"]),
        Paragraph("Sensitivity<br/>@ 100 Hz", styles["CellHead"]),
        Paragraph("% Tolerance", styles["CellHead"]),
        Paragraph("% Deviation", styles["CellHead"]),
    ]
    equip_data = [equip_header]
    equip_data.append([
        Paragraph("PVC", styles["CellText"]),
        Paragraph(info.get("pvc_model", ""), styles["CellText"]),
        Paragraph(info.get("pvc_serial", ""), styles["CellText"]),
        Paragraph(info.get("pvc_cal_date", ""), styles["CellText"]),
        Paragraph(f"{info.get('pvc_sensitivity', '')} {info.get('pvc_sens_unit', '')}", styles["CellText"]),
        Paragraph(info.get("pvc_tolerance", ""), styles["CellText"]),
        Paragraph(info.get("pvc_deviation", ""), styles["CellText"]),
    ])
    equip_data.append([
        Paragraph("Sensor", styles["CellText"]),
        Paragraph(info.get("sensor_model", ""), styles["CellText"]),
        Paragraph(info.get("sensor_serial", ""), styles["CellText"]),
        Paragraph(info.get("sensor_cal_date", ""), styles["CellText"]),
        Paragraph(f"{info.get('sensor_sensitivity', '')} {info.get('sensor_sens_unit', '')}", styles["CellText"]),
        Paragraph(info.get("sensor_tolerance", ""), styles["CellText"]),
        Paragraph(info.get("sensor_deviation", ""), styles["CellText"]),
    ])
    et = Table(equip_data, colWidths=equip_cols)
    et.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), WEARCHECK_BLACK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
    ]))
    elements.append(et)
    elements.append(Spacer(1, 3 * mm))

    # ── Customer ──
    if info.get("customer"):
        elements.append(Paragraph(
            f"<b>Customer:</b> {info['customer']}", styles["FieldValue"],
        ))
        elements.append(Spacer(1, 2 * mm))

    # ── Data Table ──
    elements.append(build_section_header("DATA TABLE", styles))
    data_rows = info.get("data_table", [])
    col1_hdr = info.get("data_col1", "Input").replace("\n", "<br/>")
    col2_hdr = info.get("data_col2", "Measured").replace("\n", "<br/>")
    col3_hdr = info.get("data_col3", "% Deviation")

    data_total_w = PAGE_W - 2 * MARGIN
    data_cols = [data_total_w * 0.33, data_total_w * 0.33, data_total_w * 0.34]

    table_data = [[
        Paragraph(col1_hdr, styles["CellHead"]),
        Paragraph(col2_hdr, styles["CellHead"]),
        Paragraph(col3_hdr, styles["CellHead"]),
    ]]
    for row in data_rows:
        inp, meas, dev = row
        try:
            dev_val = float(dev) if dev is not None else 0
            dev_str = f"{dev_val:.2f}"
        except (ValueError, TypeError):
            dev_str = str(dev) if dev else ""
            dev_val = 0

        table_data.append([
            Paragraph(f"{inp}", styles["CellText"]),
            Paragraph(f"{meas}", styles["CellText"]),
            Paragraph(dev_str, styles["CellText"]),
        ])

    dt = Table(table_data, colWidths=data_cols)
    dt_style = [
        ("BACKGROUND", (0, 0), (-1, 0), WEARCHECK_BLACK),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT_GREY]),
    ]
    dt.setStyle(TableStyle(dt_style))
    elements.append(dt)
    elements.append(Spacer(1, 3 * mm))

    # ── Note ──
    if info.get("note"):
        elements.append(Paragraph(
            f"<b>Note:</b> {info['note']}", styles["SmallItalic"],
        ))

    # Build with header/footer
    hf = make_header_footer(info)
    doc.build(elements, onFirstPage=hf, onLaterPages=hf)
    return output_path


def merge_pdfs(cover_path, data_pdfs, output_path):
    """Merge cover page + data PDF files into one output PDF."""
    writer = PdfWriter()

    # Add cover page
    cover_reader = PdfReader(cover_path)
    for page in cover_reader.pages:
        writer.add_page(page)

    # Add data PDFs
    for pdf_path in data_pdfs:
        reader = PdfReader(pdf_path)
        for page in reader.pages:
            writer.add_page(page)

    with open(output_path, "wb") as f:
        writer.write(f)

    return output_path


# ── Find files in data folder ───────────────────────────────────────────

def find_files(data_folder):
    """Locate ACC and VEL Excel and PDF files in the data folder."""
    xlsx_files = glob.glob(os.path.join(data_folder, "*.xlsx"))
    pdf_files  = glob.glob(os.path.join(data_folder, "*.pdf"))

    acc_xlsx = vel_xlsx = acc_pdf = vel_pdf = None

    for f in xlsx_files:
        name = os.path.basename(f).lower()
        if "acc" in name:
            acc_xlsx = f
        elif "vel" in name:
            vel_xlsx = f

    for f in pdf_files:
        name = os.path.basename(f).lower()
        if "acc" in name:
            acc_pdf = f
        elif "vel" in name:
            vel_pdf = f

    return acc_xlsx, vel_xlsx, acc_pdf, vel_pdf


# ── Main ─────────────────────────────────────────────────────────────────

def main():
    # Determine input: a folder or individual file(s)
    if len(sys.argv) < 2:
        data_folder = os.path.join(SCRIPT_DIR, "Data")
        if not os.path.isdir(data_folder):
            print("Usage: python generate_full_report.py <data_folder>")
            print("       python generate_full_report.py <file1.xlsx> [file2.xlsx]")
            sys.exit(1)
        args = [data_folder]
    else:
        args = sys.argv[1:]

    # Check if first argument is a file or a folder
    acc_xlsx = vel_xlsx = acc_pdf = vel_pdf = None
    data_folder = None

    if os.path.isdir(args[0]):
        data_folder = args[0]
        print(f"Scanning folder: {data_folder}")
        acc_xlsx, vel_xlsx, acc_pdf, vel_pdf = find_files(data_folder)
    else:
        # Individual file(s) passed
        for arg in args:
            if not os.path.isfile(arg):
                print(f"ERROR: File not found: {arg}")
                sys.exit(1)
            name_lower = os.path.basename(arg).lower()
            ext = name_lower.rsplit('.', 1)[-1]
            if ext == 'xlsx':
                if 'acc' in name_lower:
                    acc_xlsx = arg
                elif 'vel' in name_lower:
                    vel_xlsx = arg
            elif ext == 'pdf':
                if 'acc' in name_lower:
                    acc_pdf = arg
                elif 'vel' in name_lower:
                    vel_pdf = arg
        # Output folder = folder of the first file
        data_folder = os.path.dirname(os.path.abspath(args[0]))

    if not acc_xlsx and not vel_xlsx:
        print("ERROR: Could not find any ACC or VEL .xlsx files")
        sys.exit(1)

    if acc_xlsx:
        print(f"  ACC Excel: {os.path.basename(acc_xlsx)}")
    if vel_xlsx:
        print(f"  VEL Excel: {os.path.basename(vel_xlsx)}")
    if acc_pdf:
        print(f"  ACC PDF:   {os.path.basename(acc_pdf)}")
    elif acc_xlsx:
        print("  WARNING: No ACC .pdf found — data page will be skipped")
    if vel_pdf:
        print(f"  VEL PDF:   {os.path.basename(vel_pdf)}")
    elif vel_xlsx:
        print("  WARNING: No VEL .pdf found — data page will be skipped")

    # Read data from available workbooks
    acc_info = vel_info = None

    if acc_xlsx:
        print("\nReading ACC workbook...")
        acc_info = read_vib_analyzer(acc_xlsx)
        print(f"  Test type: {acc_info.get('test_type', '?')}")

    if vel_xlsx:
        print("Reading VEL workbook...")
        vel_info = read_vib_analyzer(vel_xlsx)
        print(f"  Test type: {vel_info.get('test_type', '?')}")

    # Merge info for cover page
    merged = merge_info(acc_info, vel_info)
    print(f"\nAnalyzer: {merged['manufacturer']} {merged['model']}")
    print(f"Serial:   {merged['serial']}")
    print(f"Cal Date: {merged['cal_date']}")

    # Generate cover page
    cover_path = os.path.join(data_folder, "_cover_page_temp.pdf")
    print("\nGenerating cover page...")
    generate_cover_page(merged, cover_path)
    print(f"  Cover page: {cover_path}")

    # Collect data PDFs to append (generate from Excel if no PDF provided)
    data_pdfs = []
    temp_data_pdfs = []

    if acc_pdf:
        data_pdfs.append(acc_pdf)
    elif acc_info:
        print("  Rendering ACC data page from Excel...")
        acc_data_path = os.path.join(data_folder, "_acc_data_temp.pdf")
        generate_data_page(acc_info, acc_data_path)
        data_pdfs.append(acc_data_path)
        temp_data_pdfs.append(acc_data_path)

    if vel_pdf:
        data_pdfs.append(vel_pdf)
    elif vel_info:
        print("  Rendering VEL data page from Excel...")
        vel_data_path = os.path.join(data_folder, "_vel_data_temp.pdf")
        generate_data_page(vel_info, vel_data_path)
        data_pdfs.append(vel_data_path)
        temp_data_pdfs.append(vel_data_path)

    # Build output filename
    model = merged.get("model", "Unknown").replace(" ", "_").replace("/", "-")
    serial = merged.get("serial", "Unknown").replace(" ", "_")
    cal_date = merged.get("cal_date", datetime.now().strftime("%Y-%m-%d"))
    # Sanitise date for filesystem (replace / with -)
    cal_date_safe = cal_date.replace("/", "-")
    output_name = f"FullReport_{model}_{serial}_{cal_date_safe}.pdf"
    output_path = os.path.join(data_folder, output_name)

    # Merge
    print("Merging into final report...")
    merge_pdfs(cover_path, data_pdfs, output_path)
    print(f"\nFinal report saved: {output_path}")

    # Clean up temp files
    for tmp in [cover_path] + temp_data_pdfs:
        try:
            os.remove(tmp)
        except OSError:
            pass

    print("Done!")


if __name__ == "__main__":
    main()
