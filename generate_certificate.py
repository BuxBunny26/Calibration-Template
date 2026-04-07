"""
WearCheck ARC - PDF Calibration Certificate Generator
======================================================
Reads a filled Excel workbook and generates a professional
two-page PDF calibration certificate.

Page 1: Certificate summary (signable, one-page)
Page 2: Full calibration data tables with deviations
"""

import os
import sys
import io
from datetime import datetime

import qrcode

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm, cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer,
    PageBreak, HRFlowable, Frame, PageTemplate, BaseDocTemplate,
    Image
)
from reportlab.platypus.flowables import KeepTogether
from reportlab.pdfgen import canvas
from reportlab.graphics.shapes import Drawing, Rect, String, Circle, Line
from reportlab.graphics import renderPDF

import openpyxl
from datetime import date, time as dt_time

# ── Brand Colors ──────────────────────────────────────────────────────────

WEARCHECK_RED = colors.HexColor("#C2040B")
WEARCHECK_BLACK = colors.HexColor("#1A1A1A")
WEARCHECK_GREY = colors.HexColor("#4D4D4D")
WEARCHECK_LTGREY = colors.HexColor("#E6E6E6")
LIGHT_GREY = colors.HexColor("#F2F2F2")
PASS_GREEN = colors.HexColor("#C6EFCE")
FAIL_RED = colors.HexColor("#FFC7CE")
DARK_GREEN = colors.HexColor("#006100")
DARK_RED = colors.HexColor("#9C0006")

# Logo path
LOGO_PATH = os.path.join(os.path.dirname(os.path.abspath(__file__)), "WearCheck Logo.png")
NMISA_QR_URL = "https://nmisa.microsoftcrmportals.com/QRCertificates/?data=NMISA-AUV-2026-11287"

PAGE_W, PAGE_H = A4
MARGIN = 15 * mm


# ── Custom Styles ─────────────────────────────────────────────────────────

def get_styles():
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(
        name="CertTitle",
        fontName="Helvetica-Bold",
        fontSize=16,
        textColor=WEARCHECK_BLACK,
        alignment=TA_CENTER,
        spaceAfter=2 * mm,
    ))
    styles.add(ParagraphStyle(
        name="SectionHead",
        fontName="Helvetica-Bold",
        fontSize=9,
        textColor=WEARCHECK_BLACK,
        alignment=TA_LEFT,
        leftIndent=2 * mm,
        spaceBefore=1.5 * mm,
        spaceAfter=0.5 * mm,
    ))
    styles.add(ParagraphStyle(
        name="SubSectionHead",
        fontName="Helvetica-Bold",
        fontSize=8,
        textColor=colors.HexColor("#4D4D4D"),
        alignment=TA_LEFT,
        leftIndent=4 * mm,
        spaceBefore=1 * mm,
        spaceAfter=0.5 * mm,
    ))
    styles.add(ParagraphStyle(
        name="FieldLabel",
        fontName="Helvetica-Bold",
        fontSize=7.5,
        textColor=WEARCHECK_BLACK,
        alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        name="FieldValue",
        fontName="Helvetica",
        fontSize=8,
        alignment=TA_LEFT,
    ))
    styles.add(ParagraphStyle(
        name="SmallItalic",
        fontName="Helvetica-Oblique",
        fontSize=6.5,
        textColor=colors.grey,
        alignment=TA_JUSTIFY,
        spaceBefore=1 * mm,
    ))
    styles.add(ParagraphStyle(
        name="ResultPass",
        fontName="Helvetica-Bold",
        fontSize=10,
        textColor=DARK_GREEN,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="ResultFail",
        fontName="Helvetica-Bold",
        fontSize=10,
        textColor=DARK_RED,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="Footer",
        fontName="Helvetica",
        fontSize=6,
        textColor=colors.grey,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="TableHeader",
        fontName="Helvetica-Bold",
        fontSize=7,
        textColor=colors.white,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="TableCell",
        fontName="Helvetica",
        fontSize=7,
        alignment=TA_CENTER,
    ))
    styles.add(ParagraphStyle(
        name="TableCellBold",
        fontName="Helvetica-Bold",
        fontSize=7,
        alignment=TA_CENTER,
    ))
    return styles


# ── Excel Data Reader ─────────────────────────────────────────────────────

class CertificateData:
    """Read all data from the filled Excel workbook."""

    def __init__(self, excel_path):
        self.wb = openpyxl.load_workbook(excel_path, data_only=True)
        self.input_ws = self.wb["Certificate_Input"]
        self.cal_ws = self.wb["Calibration_Data"]
        self.ref_ws = self.wb["Reference_Standards"]

    def _val(self, ws, row, col):
        """Get cell value, return empty string if None."""
        v = ws.cell(row=row, column=col).value
        return str(v) if v is not None else ""

    def _date_val(self, ws, row, col):
        v = ws.cell(row=row, column=col).value
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d")
        return str(v) if v is not None else ""

    def get_certificate_info(self):
        """Extract all certificate metadata from Certificate_Input sheet."""
        ws = self.input_ws
        # Map labels from column B (left fields) and column E (right fields)
        b_labels = {}  # label -> row  (values in col C)
        e_labels = {}  # label -> row  (values in col F)
        for row in range(1, ws.max_row + 1):
            label_b = self._val(ws, row, 2).strip()
            if label_b:
                b_labels[label_b] = row
            label_e = self._val(ws, row, 5).strip()
            if label_e:
                e_labels[label_e] = row

        def get(label):
            """Get value from col C by B-label."""
            r = b_labels.get(label)
            return self._val(ws, r, 3) if r else ""

        def get_e(label):
            """Get value from col F by E-label."""
            r = e_labels.get(label)
            return self._val(ws, r, 6) if r else ""

        def get_date(label):
            r = b_labels.get(label)
            return self._date_val(ws, r, 3) if r else ""

        def get_date_e(label):
            """Get date from col F by E-label."""
            r = e_labels.get(label)
            return self._date_val(ws, r, 6) if r else ""

        return {
            "cert_number": get("Certificate Number"),
            "date_of_issue": get_date("Date of Issue"),
            "calibration_date": get_date("Calibration Date"),
            "next_cal_due": get_date("Next Calibration Due"),
            "procedure": get("Calibration Procedure"),
            "equipment_name": get("Equipment Name"),
            "manufacturer": get("Manufacturer"),
            "model": get("Model Number"),
            "serial": get("Serial Number"),
            "asset": get("Asset Number"),
            "location": get("Location / Department"),
            "det_tolerance": get("Detector Tolerance (+/-)"),
            "inc_tolerance": get("Inclinometer Tolerance (+/-)"),
            "gyro_tolerance": get("Gyroscope Tolerance (+/-)"),
            "temperature": get("Temperature (°C)"),
            "humidity": get("Humidity (%RH)"),
            "pressure": get("Atmospheric Pressure (kPa)"),
            "sw_manufacturer": get("Software Manufacturer"),
            "sw_name": get("Software Name"),
            "sw_version": get("Software Version"),
            "mounting_torque": get("Mounting Torque"),
            "lubrication": get("Lubrication Used"),
            "mounting_orientation": get("Mounting Orientation"),
            "cable_routing": get("Cable Routing"),
            "reference_level": get("Reference Level"),
            "g_n": get("gn Value"),
            "tech_name": get("Tech Name"),
            "tech_surname": get_e("Tech Surname"),
            "tech_title": get("Tech Title"),
            "tech_date": get_date_e("Tech Date Signed"),
            "rev_name": get("Reviewer Name"),
            "rev_surname": get_e("Reviewer Surname"),
            "rev_title": get("Reviewer Title"),
            "rev_date": get_date_e("Reviewer Date Signed"),
            "appr_name": get("Approver Name"),
            "appr_surname": get_e("Approver Surname"),
            "appr_title": get("Approver Title"),
            "appr_date": get_date_e("Approver Date Signed"),
            "pre_cal_result": get("Pre-Calibration Result"),
            "post_cal_result": get("Post-Calibration Result"),
            "inclinometer_result": get("Inclinometer Result"),
            "gyroscope_result": get("Gyroscope Result"),
            "overall_result": get("OVERALL RESULT"),
        }

    def get_cal_table(self, start_row, num_rows=15):
        """Extract calibration table data starting at given row."""
        ws = self.cal_ws
        rows = []
        for r in range(start_row, start_row + num_rows):
            set_pt = ws.cell(row=r, column=3).value
            actual = ws.cell(row=r, column=4).value
            dev = ws.cell(row=r, column=5).value
            tol = ws.cell(row=r, column=6).value
            conf = ws.cell(row=r, column=7).value
            if set_pt is not None and set_pt != "":
                rows.append({
                    "set_point": set_pt,
                    "actual": actual,
                    "deviation": dev,
                    "tolerance": tol,
                    "conformance": conf or "",
                })
        section_result = ws.cell(row=start_row, column=8).value or ""
        return rows, str(section_result)

    def get_reference_standards(self):
        ws = self.ref_ws
        standards = []
        for row in range(4, 14):
            name = ws.cell(row=row, column=3).value
            if name:
                standards.append({
                    "name": name,
                    "manufacturer": self._val(ws, row, 4),
                    "serial": self._val(ws, row, 5),
                    "cert_number": self._val(ws, row, 6),
                    "cal_due": self._date_val(ws, row, 7),
                    "traceability": self._val(ws, row, 8),
                })
        return standards


# ── PVC Workbook Reader ───────────────────────────────────────────────────

class PVCData:
    """Read calibration data from PVC Report Generation Workbook (.xltm/.xlsm)."""

    # Cell values that indicate missing data (formula errors in blank templates)
    _BLANK = {None, "", "#N/A", "#VALUE!", "#REF!", "#DIV/0!"}

    def __init__(self, excel_path):
        self.wb = openpyxl.load_workbook(excel_path, data_only=True, keep_vba=False)
        self.has_fr = "FRCert" in self.wb.sheetnames
        self.has_lin = "LINCert" in self.wb.sheetnames
        self.has_slin = "SLINCert" in self.wb.sheetnames
        self.has_vib = "Vib Analyzer Cert" in self.wb.sheetnames

    def _val(self, ws, row, col):
        v = ws.cell(row=row, column=col).value
        if v is None or str(v).strip() in self._BLANK:
            return ""
        return str(v).strip()

    def _num(self, ws, row, col):
        v = ws.cell(row=row, column=col).value
        if v is None or str(v).strip() in self._BLANK:
            return ""
        try:
            return float(v)
        except (ValueError, TypeError):
            return str(v).strip()

    def _date_val(self, ws, row, col):
        v = ws.cell(row=row, column=col).value
        if isinstance(v, datetime):
            return v.strftime("%Y-%m-%d")
        if isinstance(v, date):
            return v.strftime("%Y-%m-%d")
        if isinstance(v, dt_time):
            if v == dt_time(0, 0):
                return ""
            return v.strftime("%H:%M")
        if v is None or str(v).strip() in self._BLANK:
            return ""
        return str(v).strip()

    def _with_unit(self, ws, val_row, val_col, unit_row, unit_col):
        val = self._val(ws, val_row, val_col)
        unit = self._val(ws, unit_row, unit_col)
        if val and unit:
            return f"{val} {unit}"
        return val

    def get_certificate_info(self):
        """Extract certificate metadata — returns dict compatible with PDF builder."""
        info = {"source": "pvc"}

        if self.has_fr:
            ws = self.wb["FRCert"]
            info.update({
                "equipment_name": self._val(ws, 9, 3),       # C9: Description
                "model": self._val(ws, 5, 3),                # C5: Model Number
                "serial": self._val(ws, 6, 3),               # C6: Serial Number
                "manufacturer": self._val(ws, 7, 3),         # C7: Manufacturer
                "asset": self._val(ws, 8, 3),                # C8: ID Number
                "sensitivity_value": self._val(ws, 17, 8),   # H17: Sensitivity @ ref
                "sensitivity_unit": self._val(ws, 17, 6),    # F17: Unit
                "ref_frequency": self._val(ws, 19, 6),       # F19: Ref frequency
                "ref_freq_unit": self._val(ws, 19, 8),       # H19: Hz
                "test_level": self._val(ws, 6, 7),           # G6: Test level
                "amp_range": self._with_unit(ws, 5, 11, 5, 12),   # K5 + L5
                "resolution": self._with_unit(ws, 6, 11, 6, 12),  # K6 + L6
                "resonant_freq": self._with_unit(ws, 7, 11, 7, 12),  # K7 + L7
                "temp_range_spec": self._with_unit(ws, 8, 11, 8, 12),  # K8 + L8
                "customer": self._val(ws, 41, 6),            # F41
                "as_found": self._val(ws, 49, 6),            # F49
                "as_left": self._val(ws, 50, 6),             # F50
                "temperature": self._val(ws, 49, 11),        # K49
                "humidity": self._val(ws, 50, 11),           # K50
                "pvc_info": self._val(ws, 31, 5),            # E31: PVC M/N, S/N, FW
                "iso_method": self._val(ws, 32, 5),          # E32: ISO method
                "user_notes": self._val(ws, 37, 5),          # F37+: user notes
                "tech_name": self._val(ws, 55, 6),           # F55: Technician
                "appr_name": self._val(ws, 56, 6),           # F56: Approval
                "calibration_date": self._date_val(ws, 55, 11),  # K55
                "calibration_time": self._date_val(ws, 56, 11),  # K56
                "next_cal_due": self._date_val(ws, 57, 11),      # K57
            })
        elif self.has_lin:
            ws = self.wb["LINCert"]
            info.update({
                "equipment_name": "",
                "model": self._val(ws, 5, 5),                # E5: Model
                "serial": self._val(ws, 6, 5),               # E6: Serial
                "manufacturer": self._val(ws, 7, 5),         # E7: Manufacturer
                "asset": self._val(ws, 8, 5),                # E8: ID Number
                "sensitivity_value": self._val(ws, 5, 15),   # O5
                "sensitivity_unit": self._val(ws, 5, 18),    # R5
                "test_frequency": self._val(ws, 7, 15),      # O7
                "max_linearity": self._val(ws, 9, 15),       # O9
                "amp_range": self._with_unit(ws, 5, 25, 5, 28),  # Y5 + AB5
                "resolution": self._with_unit(ws, 6, 25, 6, 28),
                "resonant_freq": self._with_unit(ws, 7, 25, 7, 28),
                "temp_range_spec": self._with_unit(ws, 8, 25, 8, 28),
                "customer": self._val(ws, 41, 12),           # L41
                "temperature": self._val(ws, 48, 25),        # Y48
                "humidity": self._val(ws, 50, 25),           # Y50
                "tech_name": self._val(ws, 55, 15),          # O55
                "appr_name": self._val(ws, 56, 15),          # O56
                "calibration_date": self._date_val(ws, 55, 25),  # Y55
                "next_cal_due": self._date_val(ws, 57, 25),      # Y57
            })

        # Defaults for fields the PDF builder expects but PVC doesn't have
        defaults = {
            "cert_number": "", "date_of_issue": "", "location": "",
            "procedure": "Back-to-Back Comparison per ISO 16063-21",
            "det_tolerance": "", "inc_tolerance": "", "gyro_tolerance": "",
            "pressure": "",
            "sw_manufacturer": "", "sw_name": "", "sw_version": "",
            "mounting_torque": "", "lubrication": "", "mounting_orientation": "",
            "cable_routing": "", "reference_level": "", "g_n": "",
            "tech_surname": "", "tech_title": "", "tech_date": "",
            "rev_name": "", "rev_surname": "", "rev_title": "", "rev_date": "",
            "appr_surname": "", "appr_title": "", "appr_date": "",
            "pre_cal_result": "", "post_cal_result": "",
            "inclinometer_result": "", "gyroscope_result": "",
            "overall_result": "", "equipment_name": "",
            "manufacturer": "", "model": "", "serial": "", "asset": "",
            "temperature": "", "humidity": "", "customer": "",
            "calibration_date": "", "next_cal_due": "",
        }
        for k, v in defaults.items():
            info.setdefault(k, v)

        # Generate a cert number if not available
        if not info["cert_number"]:
            sn = info.get("serial", "UNKNOWN")
            dt = info.get("calibration_date", "")
            info["cert_number"] = f"PVC-{sn}-{dt}" if dt else f"PVC-{sn}"

        if not info["date_of_issue"]:
            info["date_of_issue"] = datetime.now().strftime("%Y-%m-%d")

        return info

    def get_fr_data(self):
        """Get frequency response data from FRData sheet."""
        if "FRData" not in self.wb.sheetnames:
            return []
        ws = self.wb["FRData"]
        ref_freq = self._num(ws, 8, 2)  # B8: reference frequency
        rows = []
        for r in range(13, 46):
            freq = self._num(ws, r, 1)   # A: Frequency
            amp = self._num(ws, r, 2)    # B: Amplitude
            sens = self._num(ws, r, 3)   # C: Sensitivity
            dev = self._num(ws, r, 4)    # D: Deviation
            if freq != "":
                rows.append({
                    "frequency": freq,
                    "amplitude": amp,
                    "sensitivity": sens,
                    "deviation": dev,
                })
        return rows, ref_freq

    def get_lin_data(self):
        """Get dynamic linearity data from LINData sheet."""
        if "LINData" not in self.wb.sheetnames:
            return []
        ws = self.wb["LINData"]
        rows = []
        for r in range(14, 44):
            inp = self._num(ws, r, 2)    # B: Input
            out = self._num(ws, r, 3)    # C: Output
            sens = self._num(ws, r, 4)   # D: Sensitivity
            lin = self._num(ws, r, 5)    # E: Linearity
            if inp != "":
                rows.append({
                    "input": inp,
                    "output": out,
                    "sensitivity": sens,
                    "linearity": lin,
                })
        return rows

    def get_cal_table(self, start_row, num_rows=15):
        """Not applicable for PVC — return empty."""
        return [], ""

    def get_reference_standards(self):
        """PVC is the reference standard — parsed from E31 text."""
        return []


# ── PDF Builder ───────────────────────────────────────────────────────────

def make_header_footer(info):
    """Return a header/footer draw function with access to certificate info."""
    def header_footer(canvas_obj, doc):
        """Draw header and footer on every page."""
        canvas_obj.saveState()

        # Header — clean white background

        # Logo — pinned to top-left corner
        if os.path.exists(LOGO_PATH):
            try:
                logo_h = 20 * mm
                logo_w = logo_h * 488 / 511
                canvas_obj.drawImage(
                    LOGO_PATH,
                    MARGIN, PAGE_H - MARGIN - logo_h,
                    width=logo_w, height=logo_h,
                    preserveAspectRatio=True,
                    mask='auto'
                )
            except Exception:
                canvas_obj.setFillColor(WEARCHECK_BLACK)
                canvas_obj.setFont("Helvetica-Bold", 14)
                canvas_obj.drawString(MARGIN, PAGE_H - 11 * mm, "WEARCHECK")
        else:
            canvas_obj.setFillColor(WEARCHECK_BLACK)
            canvas_obj.setFont("Helvetica-Bold", 14)
            canvas_obj.drawString(MARGIN, PAGE_H - 11 * mm, "WEARCHECK")

        # Header text — right side
        canvas_obj.setFillColor(WEARCHECK_BLACK)
        canvas_obj.setFont("Helvetica-Bold", 11)
        canvas_obj.drawRightString(PAGE_W - MARGIN, PAGE_H - 8 * mm, "CALIBRATION CERTIFICATE")
        canvas_obj.setFont("Helvetica", 7)
        canvas_obj.setFillColor(WEARCHECK_GREY)
        canvas_obj.drawRightString(PAGE_W - MARGIN, PAGE_H - 11.5 * mm, "WearCheck ARC \u2014 Condition Monitoring Division")

        # Accreditation — small text in header, right-aligned
        canvas_obj.setFont("Helvetica", 6.5)
        canvas_obj.setFillColor(WEARCHECK_BLACK)
        canvas_obj.drawRightString(
            PAGE_W - MARGIN, PAGE_H - 14.5 * mm,
            "ISO 16063-21  |  NMISA-AUV-2026-11287"
        )

        # QR code — between logo and header text
        try:
            qr = qrcode.make(NMISA_QR_URL, box_size=10, border=0)
            qr_buf = io.BytesIO()
            qr.save(qr_buf, format='PNG')
            qr_buf.seek(0)
            qr_size = 14 * mm
            qr_x = MARGIN + 22 * mm
            qr_y = PAGE_H - MARGIN - qr_size
            from reportlab.lib.utils import ImageReader
            canvas_obj.drawImage(ImageReader(qr_buf), qr_x, qr_y,
                                 width=qr_size, height=qr_size)
            canvas_obj.setFont("Helvetica", 4.5)
            canvas_obj.setFillColor(WEARCHECK_GREY)
            canvas_obj.drawCentredString(qr_x + qr_size / 2, qr_y - 3 * mm,
                                         "Scan to authenticate")
        except Exception:
            pass

        # Footer line — subtle grey
        canvas_obj.setStrokeColor(colors.lightgrey)
        canvas_obj.setLineWidth(0.5)
        canvas_obj.line(MARGIN, 14 * mm, PAGE_W - MARGIN, 14 * mm)

        # Footer — equipment details left, generation info right
        canvas_obj.setFillColor(WEARCHECK_BLACK)
        canvas_obj.setFont("Helvetica", 6.5)
        equip_line = (
            f"{info.get('equipment_name', '')}  |  "
            f"S/N: {info.get('serial', '')}  |  "
            f"Cert: {info.get('cert_number', '')}  |  "
            f"Cal Date: {info.get('calibration_date', '')}"
        )
        canvas_obj.drawString(MARGIN, 10 * mm, equip_line)
        canvas_obj.drawRightString(
            PAGE_W - MARGIN, 10 * mm,
            f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}  |  Page {doc.page}"
        )

        canvas_obj.setFont("Helvetica", 6)
        canvas_obj.setFillColor(WEARCHECK_GREY)
        canvas_obj.drawString(
            MARGIN, 6.5 * mm,
            "This certificate is issued in accordance with WearCheck ARC quality management procedures aligned with ISO/IEC 17025."
        )

        canvas_obj.restoreState()
    return header_footer


def build_section_header(text, styles):
    """Section header — print-friendly."""
    return Table(
        [[Paragraph(text, styles["SectionHead"])]],
        colWidths=[PAGE_W - 2 * MARGIN],
        rowHeights=[6 * mm],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), LIGHT_GREY),
            ("LINEBELOW", (0, 0), (-1, -1), 0.5, colors.lightgrey),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ])
    )


def build_sub_section_header(text, styles):
    """Sub-section header — lighter style to show hierarchy."""
    return Table(
        [[Paragraph(text, styles["SubSectionHead"])]],
        colWidths=[PAGE_W - 2 * MARGIN],
        rowHeights=[5 * mm],
        style=TableStyle([
            ("BACKGROUND", (0, 0), (-1, -1), colors.HexColor("#F5F5F5")),
            ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.HexColor("#DDDDDD")),
            ("TOPPADDING", (0, 0), (-1, -1), 1),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 1),
        ])
    )


def build_field_table(fields, styles, col_widths=None):
    """Two-column field: label | value pairs.
    fields: list of (label, value) tuples
    """
    if not col_widths:
        half = (PAGE_W - 2 * MARGIN) / 2
        col_widths = [half * 0.4, half * 0.6, half * 0.4, half * 0.6]

    rows = []
    for i in range(0, len(fields), 2):
        row = []
        # Left pair
        row.append(Paragraph(fields[i][0], styles["FieldLabel"]))
        row.append(Paragraph(str(fields[i][1]), styles["FieldValue"]))
        # Right pair
        if i + 1 < len(fields):
            row.append(Paragraph(fields[i + 1][0], styles["FieldLabel"]))
            row.append(Paragraph(str(fields[i + 1][1]), styles["FieldValue"]))
        else:
            row.extend(["", ""])
        rows.append(row)

    t = Table(rows, colWidths=col_widths)
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        ("LINEBELOW", (0, 0), (-1, -1), 0.3, colors.lightgrey),
        ("LEFTPADDING", (0, 0), (-1, -1), 2),
    ]))
    return t


def build_result_badge(result, styles):
    """Colored PASS/FAIL badge."""
    if result == "PASS":
        style = ParagraphStyle("pass_badge", parent=styles["ResultPass"])
        bg = PASS_GREEN
        tc = DARK_GREEN
    elif result == "FAIL":
        style = ParagraphStyle("fail_badge", parent=styles["ResultFail"])
        bg = FAIL_RED
        tc = DARK_RED
    else:
        style = ParagraphStyle("na_badge", parent=styles["FieldValue"], alignment=TA_CENTER)
        bg = LIGHT_GREY
        tc = colors.grey

    t = Table(
        [[Paragraph(str(result), style)]],
        colWidths=[30 * mm],
        rowHeights=[8 * mm],
    )
    t.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (0, 0), bg),
        ("ALIGN", (0, 0), (0, 0), "CENTER"),
        ("VALIGN", (0, 0), (0, 0), "MIDDLE"),
        ("TOPPADDING", (0, 0), (0, 0), 1),
        ("BOTTOMPADDING", (0, 0), (0, 0), 1),
        ("ROUNDEDCORNERS", [2, 2, 2, 2]),
    ]))
    return t


def build_cal_data_table(data_rows, title, styles):
    """Build a calibration data table for page 2."""
    header = [
        Paragraph("Set Point", styles["TableHeader"]),
        Paragraph("Actual", styles["TableHeader"]),
        Paragraph("Deviation", styles["TableHeader"]),
        Paragraph("Tolerance", styles["TableHeader"]),
        Paragraph("Conformance", styles["TableHeader"]),
    ]

    table_data = [header]
    for row in data_rows:
        conf_text = str(row["conformance"])
        if conf_text == "PASS":
            conf_p = Paragraph(f'<font color="#006100"><b>{conf_text}</b></font>', styles["TableCell"])
        elif conf_text == "FAIL":
            conf_p = Paragraph(f'<font color="#9C0006"><b>{conf_text}</b></font>', styles["TableCell"])
        else:
            conf_p = Paragraph(conf_text, styles["TableCell"])

        def fmt(v):
            if v is None:
                return ""
            try:
                return f"{float(v):.3f}"
            except (ValueError, TypeError):
                return str(v)

        table_data.append([
            Paragraph(fmt(row["set_point"]), styles["TableCell"]),
            Paragraph(fmt(row["actual"]), styles["TableCell"]),
            Paragraph(fmt(row["deviation"]), styles["TableCell"]),
            Paragraph(fmt(row["tolerance"]), styles["TableCell"]),
            conf_p,
        ])

    col_w = (PAGE_W - 2 * MARGIN) / 5
    t = Table(table_data, colWidths=[col_w] * 5, repeatRows=1)

    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), WEARCHECK_GREY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
    ]

    # Alternate row shading
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), LIGHT_GREY))

        # Color conformance cell background
        if i < len(table_data):
            conf_val = data_rows[i - 1]["conformance"] if i - 1 < len(data_rows) else ""
            if conf_val == "PASS":
                style_cmds.append(("BACKGROUND", (4, i), (4, i), PASS_GREEN))
            elif conf_val == "FAIL":
                style_cmds.append(("BACKGROUND", (4, i), (4, i), FAIL_RED))

    t.setStyle(TableStyle(style_cmds))
    return t


def build_signature_table(info, styles):
    """Flat authorisation table with gap columns so lines don't merge."""
    sig_lbl = ParagraphStyle("sig_lbl", parent=styles["SmallItalic"], alignment=TA_LEFT)
    total_w = PAGE_W - 2 * MARGIN
    gap = 4 * mm
    block_w = (total_w - 2 * gap) / 3
    name_w = block_w * 0.55
    title_w = block_w * 0.45
    col_widths = [name_w, title_w, gap, name_w, title_w, gap, name_w, title_w]

    E = ""  # empty cell for gap columns

    sig_data = [
        # Row 0: Role headers (span across each pair)
        [Paragraph("<b>Calibrated By:</b>", styles["FieldLabel"]), "", E,
         Paragraph("<b>Reviewed By:</b>", styles["FieldLabel"]), "", E,
         Paragraph("<b>Approved By:</b>", styles["FieldLabel"]), ""],
        # Row 1: Name + Surname | Title
        [Paragraph(f"{info['tech_name']} {info['tech_surname']}", styles["FieldValue"]),
         Paragraph(str(info['tech_title']), styles["FieldValue"]), E,
         Paragraph(f"{info['rev_name']} {info['rev_surname']}", styles["FieldValue"]),
         Paragraph(str(info['rev_title']), styles["FieldValue"]), E,
         Paragraph(f"{info['appr_name']} {info['appr_surname']}", styles["FieldValue"]),
         Paragraph(str(info['appr_title']), styles["FieldValue"])],
        # Row 2: Sub-labels
        [Paragraph("Name & Surname", sig_lbl), Paragraph("Title / Designation", sig_lbl), E,
         Paragraph("Name & Surname", sig_lbl), Paragraph("Title / Designation", sig_lbl), E,
         Paragraph("Name & Surname", sig_lbl), Paragraph("Title / Designation", sig_lbl)],
        # Row 3: spacer for signature area
        ["", "", E, "", "", E, "", ""],
        # Row 4: Signature | Date Signed
        ["", Paragraph(str(info['tech_date']), styles["FieldValue"]), E,
         "", Paragraph(str(info['rev_date']), styles["FieldValue"]), E,
         "", Paragraph(str(info['appr_date']), styles["FieldValue"])],
        # Row 5: Sub-labels
        [Paragraph("Signature", sig_lbl), Paragraph("Date Signed", sig_lbl), E,
         Paragraph("Signature", sig_lbl), Paragraph("Date Signed", sig_lbl), E,
         Paragraph("Signature", sig_lbl), Paragraph("Date Signed", sig_lbl)],
    ]

    t = Table(sig_data, colWidths=col_widths, rowHeights=[None, None, None, 8 * mm, None, None])
    t.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "BOTTOM"),
        ("TOPPADDING", (0, 0), (-1, -1), 1.5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 1.5),
        # Span role headers across each pair (skip gap cols)
        ("SPAN", (0, 0), (1, 0)),
        ("SPAN", (3, 0), (4, 0)),
        ("SPAN", (6, 0), (7, 0)),
        # Lines under name + title (row 1) — per block
        ("LINEBELOW", (0, 1), (0, 1), 0.5, colors.black),
        ("LINEBELOW", (1, 1), (1, 1), 0.5, colors.black),
        ("LINEBELOW", (3, 1), (3, 1), 0.5, colors.black),
        ("LINEBELOW", (4, 1), (4, 1), 0.5, colors.black),
        ("LINEBELOW", (6, 1), (6, 1), 0.5, colors.black),
        ("LINEBELOW", (7, 1), (7, 1), 0.5, colors.black),
        # Lines under signature + date (row 4) — per block
        ("LINEBELOW", (0, 4), (0, 4), 0.5, colors.black),
        ("LINEBELOW", (1, 4), (1, 4), 0.5, colors.black),
        ("LINEBELOW", (3, 4), (3, 4), 0.5, colors.black),
        ("LINEBELOW", (4, 4), (4, 4), 0.5, colors.black),
        ("LINEBELOW", (6, 4), (6, 4), 0.5, colors.black),
        ("LINEBELOW", (7, 4), (7, 4), 0.5, colors.black),
    ]))
    return t


def build_fr_data_table(data_rows, styles):
    """Build a frequency response data table."""
    header = [
        Paragraph("Frequency", styles["TableHeader"]),
        Paragraph("Amplitude", styles["TableHeader"]),
        Paragraph("Sensitivity", styles["TableHeader"]),
        Paragraph("Deviation (%)", styles["TableHeader"]),
    ]
    table_data = [header]

    def fmt(v):
        if v == "" or v is None:
            return ""
        try:
            f = float(v)
            return f"{f:.4f}" if abs(f) < 10 else f"{f:.2f}"
        except (ValueError, TypeError):
            return str(v)

    for row in data_rows:
        table_data.append([
            Paragraph(fmt(row["frequency"]), styles["TableCell"]),
            Paragraph(fmt(row["amplitude"]), styles["TableCell"]),
            Paragraph(fmt(row["sensitivity"]), styles["TableCell"]),
            Paragraph(fmt(row["deviation"]), styles["TableCell"]),
        ])

    col_w = (PAGE_W - 2 * MARGIN) / 4
    t = Table(table_data, colWidths=[col_w] * 4, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), WEARCHECK_GREY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
    ]
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), LIGHT_GREY))
    t.setStyle(TableStyle(style_cmds))
    return t


def build_lin_data_table(data_rows, styles):
    """Build a linearity data table."""
    header = [
        Paragraph("Input", styles["TableHeader"]),
        Paragraph("Output", styles["TableHeader"]),
        Paragraph("Sensitivity", styles["TableHeader"]),
        Paragraph("Linearity (%)", styles["TableHeader"]),
    ]
    table_data = [header]

    def fmt(v):
        if v == "" or v is None:
            return ""
        try:
            f = float(v)
            return f"{f:.4f}" if abs(f) < 10 else f"{f:.2f}"
        except (ValueError, TypeError):
            return str(v)

    for row in data_rows:
        table_data.append([
            Paragraph(fmt(row["input"]), styles["TableCell"]),
            Paragraph(fmt(row["output"]), styles["TableCell"]),
            Paragraph(fmt(row["sensitivity"]), styles["TableCell"]),
            Paragraph(fmt(row["linearity"]), styles["TableCell"]),
        ])

    col_w = (PAGE_W - 2 * MARGIN) / 4
    t = Table(table_data, colWidths=[col_w] * 4, repeatRows=1)
    style_cmds = [
        ("BACKGROUND", (0, 0), (-1, 0), WEARCHECK_GREY),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("TOPPADDING", (0, 0), (-1, -1), 2),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
    ]
    for i in range(1, len(table_data)):
        if i % 2 == 0:
            style_cmds.append(("BACKGROUND", (0, i), (-1, i), LIGHT_GREY))
    t.setStyle(TableStyle(style_cmds))
    return t


def generate_certificate(excel_path, output_path=None):
    """Main PDF generation function. Auto-detects WearCheck ARC or PVC workbook format."""
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found: {excel_path}")
        sys.exit(1)

    # Auto-detect workbook format
    wb_check = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)
    sheet_names = wb_check.sheetnames
    wb_check.close()

    is_pvc = "FRCert" in sheet_names or "FRData" in sheet_names or "LINCert" in sheet_names
    if is_pvc:
        data = PVCData(excel_path)
        print("  Detected PVC Report Generation Workbook format")
    else:
        data = CertificateData(excel_path)

    info = data.get_certificate_info()
    styles = get_styles()

    if output_path is None:
        cert_num = info["cert_number"].replace("/", "-").replace("\\", "-")
        if not cert_num or cert_num == "None":
            cert_num = "DRAFT"
        output_path = os.path.join(
            os.path.dirname(excel_path),
            f"Certificate_{cert_num}.pdf"
        )

    doc = SimpleDocTemplate(
        output_path,
        pagesize=A4,
        topMargin=20 * mm,
        bottomMargin=15 * mm,
        leftMargin=MARGIN,
        rightMargin=MARGIN,
        title=f"Calibration Certificate {info.get('cert_number', '')}",
        author="WearCheck ARC",
    )

    elements = []

    # ══════════════════════════════════════════════════════════════════════
    # PAGE 1: CERTIFICATE SUMMARY
    # ══════════════════════════════════════════════════════════════════════

    # Certificate title
    elements.append(Spacer(1, 1 * mm))
    elements.append(Paragraph("CERTIFICATE OF CALIBRATION", styles["CertTitle"]))
    elements.append(HRFlowable(
        width="100%", thickness=0.5, color=colors.lightgrey, spaceAfter=1.5 * mm
    ))

    # Certificate info
    elements.append(build_section_header("CERTIFICATE INFORMATION", styles))
    cert_info_fields = [
        ("Certificate No:", info["cert_number"]),
        ("Date of Issue:", info["date_of_issue"]),
        ("Calibration Date:", info["calibration_date"]),
        ("Next Due:", info["next_cal_due"]),
    ]
    if not is_pvc:
        cert_info_fields += [("Procedure:", info["procedure"]), ("", "")]
    else:
        cert_info_fields += [
            ("Customer:", info.get("customer", "")),
            ("", ""),
        ]
    elements.append(build_field_table(cert_info_fields, styles))

    if is_pvc:
        # ── PVC-specific page 1 ──────────────────────────────────────────

        # Sensor Information
        elements.append(build_section_header("SENSOR INFORMATION", styles))
        sensor_fields = [
            ("Description:", info["equipment_name"]),
            ("Manufacturer:", info["manufacturer"]),
            ("Model:", info["model"]),
            ("Serial No:", info["serial"]),
        ]
        if info.get("asset"):
            sensor_fields += [("ID Number:", info["asset"]), ("", "")]
        elements.append(build_field_table(sensor_fields, styles))

        # Transducer Specifications
        spec_fields = [
            ("Amp. Range:", info.get("amp_range", "")),
            ("Resolution:", info.get("resolution", "")),
            ("Resonant Freq:", info.get("resonant_freq", "")),
            ("Temp. Range:", info.get("temp_range_spec", "")),
        ]
        if any(v for _, v in spec_fields):
            elements.append(build_section_header("TRANSDUCER SPECIFICATIONS", styles))
            elements.append(build_field_table(spec_fields, styles))

        # Calibration Data Summary
        elements.append(build_section_header("CALIBRATION DATA", styles))
        cal_summary = []
        if info.get("sensitivity_value"):
            sens_str = info["sensitivity_value"]
            if info.get("sensitivity_unit"):
                sens_str = f"{sens_str} {info['sensitivity_unit']}"
            cal_summary += [("Sensitivity:", sens_str)]
        else:
            cal_summary += [("Sensitivity:", "")]
        if info.get("ref_frequency"):
            rf = info["ref_frequency"]
            if info.get("ref_freq_unit"):
                rf = f"{rf} {info['ref_freq_unit']}"
            cal_summary += [("Ref. Frequency:", rf)]
        else:
            cal_summary += [("Ref. Frequency:", "")]
        if info.get("test_level"):
            cal_summary += [("Test Level:", info["test_level"]), ("", "")]
        if info.get("max_linearity"):
            cal_summary += [("Max Linearity:", f"{info['max_linearity']}%"), ("", "")]
        if len(cal_summary) % 2 != 0:
            cal_summary.append(("", ""))
        elements.append(build_field_table(cal_summary, styles))

        # Unit Condition & Lab Conditions
        elements.append(build_section_header("CONDITIONS", styles))
        cond_fields = []
        if info.get("as_found"):
            cond_fields += [("As Found:", info["as_found"])]
        if info.get("as_left"):
            cond_fields += [("As Left:", info["as_left"])]
        if info.get("temperature"):
            cond_fields += [("Temperature:", f"{info['temperature']} °C")]
        if info.get("humidity"):
            cond_fields += [("Humidity:", f"{info['humidity']} %")]
        if not cond_fields:
            cond_fields = [("Temperature:", ""), ("Humidity:", "")]
        if len(cond_fields) % 2 != 0:
            cond_fields.append(("", ""))
        elements.append(build_field_table(cond_fields, styles))

        # Procedure
        elements.append(build_section_header("PROCEDURE", styles))
        proc_text = info.get("iso_method", "") or "Back-to-Back Comparison per ISO 16063-21"
        pvc_ref = info.get("pvc_info", "")
        if pvc_ref:
            proc_text = f"{pvc_ref}<br/>{proc_text}"
        elements.append(Paragraph(proc_text, styles["FieldValue"]))

        # Traceability & notes
        elements.append(Paragraph(
            "<b>Traceability:</b> The measurements reported herein are traceable to NIST (USA) and PTB (Germany) "
            "through an unbroken chain of calibrations.",
            styles["SmallItalic"]
        ))
        elements.append(Paragraph(
            "<b>Uncertainty:</b> The reported expanded uncertainty is based on a standard uncertainty "
            "multiplied by a coverage factor k=2, providing a level of confidence of approximately 95%.",
            styles["SmallItalic"]
        ))
        if info.get("user_notes"):
            elements.append(Spacer(1, 1 * mm))
            elements.append(Paragraph(f"<b>Notes:</b> {info['user_notes']}", styles["SmallItalic"]))

        elements.append(Spacer(1, 1 * mm))

        # Authorisation (PVC has Technician + Approval, no Reviewer)
        elements.append(build_section_header("AUTHORISATION", styles))
        elements.append(build_signature_table(info, styles))

        # ══════════════════════════════════════════════════════════════════
        # PAGE 2: PVC CALIBRATION DATA TABLES
        # ══════════════════════════════════════════════════════════════════

        elements.append(PageBreak())
        elements.append(Spacer(1, 2 * mm))
        elements.append(Paragraph("CALIBRATION DATA — DETAILED RESULTS", styles["CertTitle"]))
        elements.append(HRFlowable(
            width="100%", thickness=0.5, color=colors.lightgrey, spaceAfter=3 * mm
        ))
        elements.append(build_field_table([
            ("Certificate No:", info["cert_number"]),
            ("Sensor:", f"{info['manufacturer']} {info['model']}"),
            ("Serial No:", info["serial"]),
            ("Calibration Date:", info["calibration_date"]),
        ], styles))
        elements.append(Spacer(1, 2 * mm))

        # Frequency Response table
        if data.has_fr:
            result = data.get_fr_data()
            if isinstance(result, tuple):
                fr_rows, ref_freq = result
            else:
                fr_rows, ref_freq = result, ""
            if fr_rows:
                title_text = "FREQUENCY RESPONSE"
                if ref_freq:
                    title_text += f"  —  Reference Frequency: {ref_freq} Hz"
                elements.append(build_section_header(title_text, styles))
                elements.append(Spacer(1, 1 * mm))
                elements.append(build_fr_data_table(fr_rows, styles))
                elements.append(Spacer(1, 3 * mm))
            else:
                elements.append(build_section_header("FREQUENCY RESPONSE  —  NO DATA", styles))
                elements.append(Spacer(1, 2 * mm))

        # Linearity table
        if data.has_lin:
            lin_rows = data.get_lin_data()
            if lin_rows:
                elements.append(build_section_header("DYNAMIC LINEARITY", styles))
                elements.append(Spacer(1, 1 * mm))
                elements.append(build_lin_data_table(lin_rows, styles))
                elements.append(Spacer(1, 3 * mm))
            else:
                elements.append(build_section_header("DYNAMIC LINEARITY  —  NO DATA", styles))
                elements.append(Spacer(1, 2 * mm))

        # End of data
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey, spaceAfter=2 * mm))
        elements.append(Paragraph(
            "— End of Calibration Data —",
            ParagraphStyle("endmark", parent=styles["SmallItalic"], alignment=TA_CENTER, fontSize=8)
        ))

    else:
        # ── WearCheck ARC page 1 (original) ──────────────────────────────

        # Equipment details
        elements.append(build_section_header("EQUIPMENT UNDER TEST", styles))
        elements.append(build_field_table([
            ("Equipment:", info["equipment_name"]),
            ("Manufacturer:", info["manufacturer"]),
            ("Model:", info["model"]),
            ("Serial No:", info["serial"]),
            ("Asset No:", info["asset"]),
            ("Location:", info["location"]),
        ], styles))

        # Specifications
        elements.append(build_section_header("SPECIFICATIONS & TOLERANCES", styles))
        elements.append(build_field_table([
            ("Detector Tol (+/-):", info["det_tolerance"]),
            ("Inclinometer Tol (+/-):", info["inc_tolerance"]),
            ("Gyroscope Tol (+/-):", info["gyro_tolerance"]),
            ("", ""),
        ], styles))

        # Environmental conditions
        elements.append(build_section_header("ENVIRONMENTAL CONDITIONS", styles))
        elements.append(build_field_table([
            ("Temperature:", f"{info['temperature']} °C"),
            ("Humidity:", f"{info['humidity']} %RH"),
            ("Pressure:", f"{info['pressure']} kPa"),
            ("", ""),
        ], styles))

        # Procedure
        elements.append(build_section_header("PROCEDURE", styles))
        elements.append(Paragraph(
            "The instrument was calibrated in accordance with ISO 16063-21 using reference "
            "standards traceable to NIST (USA) and PTB (Germany). The following equipment was used:",
            styles["FieldValue"]
        ))

        # Calibration Software
        if info.get("sw_name"):
            elements.append(build_sub_section_header("Calibration Software", styles))
            elements.append(build_field_table([
                ("Manufacturer:", info["sw_manufacturer"]),
                ("Name:", info["sw_name"]),
                ("Version:", info["sw_version"]),
                ("", ""),
            ], styles))

        # Mounting Conditions
        mounting_fields = [
            ("Mounting Torque:", info.get("mounting_torque", "")),
            ("Lubrication Used:", info.get("lubrication", "")),
            ("Mounting Orientation:", info.get("mounting_orientation", "")),
            ("Cable Routing:", info.get("cable_routing", "")),
            ("Reference Level:", info.get("reference_level", "")),
            ("g<sub>n</sub>:", info.get("g_n", "")),
        ]
        has_mounting = any(v for _, v in mounting_fields if v)
        if has_mounting:
            elements.append(build_sub_section_header("Mounting Conditions & Considerations", styles))
            elements.append(build_field_table(mounting_fields, styles))

        # Reference standards (condensed)
        ref_standards = data.get_reference_standards()
        if ref_standards:
            elements.append(build_sub_section_header("Reference Standards", styles))
            ref_fields = []
            for std in ref_standards[:4]:
                ref_fields.append((f"{std['name']}:", f"S/N: {std['serial']}  Cert: {std['cert_number']}"))
            if len(ref_fields) % 2 != 0:
                ref_fields.append(("", ""))
            elements.append(build_field_table(ref_fields, styles))

        # Results
        elements.append(build_section_header("CALIBRATION RESULTS", styles))

        result_data = [
            [
                Paragraph("<b>Test</b>", styles["FieldLabel"]),
                Paragraph("<b>Pre-Calibration</b>", styles["FieldLabel"]),
                Paragraph("<b>Post-Calibration</b>", styles["FieldLabel"]),
                Paragraph("<b>Inclinometer</b>", styles["FieldLabel"]),
                Paragraph("<b>Gyroscope</b>", styles["FieldLabel"]),
            ],
            [
                Paragraph("Result", styles["FieldLabel"]),
                build_result_badge(info["pre_cal_result"], styles),
                build_result_badge(info["post_cal_result"], styles),
                build_result_badge(info["inclinometer_result"], styles),
                build_result_badge(info["gyroscope_result"], styles),
            ],
        ]
        col_w = (PAGE_W - 2 * MARGIN) / 5
        result_table = Table(result_data, colWidths=[col_w] * 5)
        result_table.setStyle(TableStyle([
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("LINEBELOW", (0, 0), (-1, 0), 0.5, colors.lightgrey),
            ("TOPPADDING", (0, 0), (-1, -1), 2),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 2),
        ]))
        elements.append(result_table)

        # Overall result
        overall = info["overall_result"]
        elements.append(Spacer(1, 1 * mm))
        overall_style = styles["ResultPass"] if overall == "PASS" else styles["ResultFail"]
        overall_bg = PASS_GREEN if overall == "PASS" else FAIL_RED if overall == "FAIL" else LIGHT_GREY
        overall_table = Table(
            [[Paragraph(f"OVERALL RESULT: {overall}", overall_style)]],
            colWidths=[PAGE_W - 2 * MARGIN],
            rowHeights=[7 * mm],
        )
        overall_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (0, 0), overall_bg),
            ("ALIGN", (0, 0), (0, 0), "CENTER"),
            ("VALIGN", (0, 0), (0, 0), "MIDDLE"),
            ("LINEBELOW", (0, 0), (0, 0), 0.5, colors.lightgrey),
            ("ROUNDEDCORNERS", [2, 2, 2, 2]),
        ]))
        elements.append(overall_table)
        elements.append(Spacer(1, 1 * mm))

        # Statements
        elements.append(Paragraph(
            "<b>Traceability:</b> The measurements reported herein are traceable to NIST (USA) and PTB (Germany) "
            "through an unbroken chain of calibrations.",
            styles["SmallItalic"]
        ))
        elements.append(Paragraph(
            "<b>Uncertainty:</b> The reported expanded uncertainty is based on a standard uncertainty "
            "multiplied by a coverage factor k=2, providing a level of confidence of approximately 95%.",
            styles["SmallItalic"]
        ))
        elements.append(Spacer(1, 1 * mm))

        # Authorisation signatures
        elements.append(build_section_header("AUTHORISATION", styles))
        elements.append(build_signature_table(info, styles))

        # ══════════════════════════════════════════════════════════════════
        # PAGE 2: ARC CALIBRATION DATA
        # ══════════════════════════════════════════════════════════════════

        elements.append(PageBreak())
        elements.append(Spacer(1, 2 * mm))
        elements.append(Paragraph("CALIBRATION DATA — DETAILED RESULTS", styles["CertTitle"]))
        elements.append(HRFlowable(
            width="100%", thickness=0.5, color=colors.lightgrey, spaceAfter=3 * mm
        ))
        elements.append(build_field_table([
            ("Certificate No:", info["cert_number"]),
            ("Equipment:", info["equipment_name"]),
            ("Serial No:", info["serial"]),
            ("Calibration Date:", info["calibration_date"]),
        ], styles))
        elements.append(Spacer(1, 2 * mm))

        tables = [
            ("TABLE 1: BEFORE CALIBRATION — DETECTOR RESPONSE", 3),
            ("TABLE 2: AFTER CALIBRATION — DETECTOR RESPONSE", 21),
            ("TABLE 3: INCLINOMETER CHECK", 39),
            ("TABLE 4: GYROSCOPE CHECK", 57),
        ]

        for title, start_row in tables:
            rows, section_result = data.get_cal_table(start_row)
            if rows:
                elements.append(build_section_header(f"{title}  —  Result: {section_result}", styles))
                elements.append(Spacer(1, 1 * mm))
                elements.append(build_cal_data_table(rows, title, styles))
                elements.append(Spacer(1, 3 * mm))
            else:
                elements.append(build_section_header(f"{title}  —  NO DATA", styles))
                elements.append(Spacer(1, 2 * mm))

        # End of data statement
        elements.append(HRFlowable(width="100%", thickness=0.5, color=colors.lightgrey, spaceAfter=2 * mm))
        elements.append(Paragraph(
            "— End of Calibration Data —",
            ParagraphStyle("endmark", parent=styles["SmallItalic"], alignment=TA_CENTER, fontSize=8)
        ))

    # Build PDF
    hf = make_header_footer(info)
    doc.build(elements, onFirstPage=hf, onLaterPages=hf)
    print(f"Certificate PDF generated: {output_path}")
    return output_path


# ── CLI Entry Point ───────────────────────────────────────────────────────

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python generate_certificate.py <excel_file> [output_pdf]")
        print("Example: python generate_certificate.py WearCheck_ARC_Calibration_Template.xlsx")
        sys.exit(1)

    excel_file = sys.argv[1]
    output_pdf = sys.argv[2] if len(sys.argv) > 2 else None
    generate_certificate(excel_file, output_pdf)
